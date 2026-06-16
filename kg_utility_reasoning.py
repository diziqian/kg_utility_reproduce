# -*- coding: utf-8 -*-
from __future__ import annotations
import os, json, math, argparse
from dataclasses import dataclass
from itertools import combinations
from typing import Any, Dict, List, Tuple, Optional

import numpy as np
import pandas as pd
from sklearn.linear_model import BayesianRidge, LinearRegression
from sklearn.model_selection import KFold, GroupKFold
from sklearn.preprocessing import MultiLabelBinarizer, OneHotEncoder, StandardScaler
from sklearn.metrics import r2_score, mean_absolute_error, mean_squared_error
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.decomposition import TruncatedSVD
from sklearn.neighbors import NearestNeighbors
try:
    import networkx as nx
except Exception:  # networkx is optional; KG-reasoning outputs are skipped if unavailable
    nx = None
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

RANDOM_STATE = 42
N_SPLITS = 5
FE_MIN_LABEL_FREQ_FOR_TABLE = 3
OUTPUT_PATH = './result_kg_reproduce_mppr_v2'
EXCEL_PATH = './anymous/name_price_anonymized.xlsx'
MEDIA_PATH = './anymous/media_result.xlsx'

BLOCKS = ["Supplier FE", "Src FE", "App FE", "Product", "Supply", "Demand", "Market"]
BLOCK_SHORT = {
    "Supplier FE": "Supplier FE",
    "Src FE": "Src FE",
    "App FE": "App FE",
    "Product": "Product",
    "Supply": "Supply",
    "Demand": "Demand",
    "Market": "Market",
}


# =============================================================================
# KG reasoning controls added back into the manuscript-facing reproduction program.
# These additions are deliberately append-only: they generate the missing upstream
# reasoning/proxy files but do NOT change the block definitions, Bayesian Ridge
# estimator, table construction, CV splits, or existing outputs.
# =============================================================================
PPR_ALPHA = 0.85
PPR_TOPK_APPS = 10
PPR_TOPK_SRCS = 10
PPR_PRIOR_STRENGTH = 1.0
METAPATH_ALPHA = 0.85
METAPATH_TOPK = 10
METAPATH_PRIOR_STRENGTH_APPHEAT = 1.0
PPR_HEAT_TELEPORT_MASS = 0.35
METAPATH_HEAT_TELEPORT_MASS = 0.35
TEXT_EMB_DIM = 50
TEXT_MIN_DF = 2
SIM_KNN_K = 10
SIM_MIN_COS = 0.25
MAX_NX_PAGERANK_ITERS = 200
NX_PAGERANK_TOL = 1e-8



def ensure_dir(p: str) -> None:
    os.makedirs(p, exist_ok=True)


def write_text(path: str, text: str) -> None:
    with open(path, 'w', encoding='utf-8') as f:
        f.write(text)


def read_json(path: str) -> Any:
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_json(path: str, obj: Any) -> None:
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)


def parse_asset_num(x: Any) -> float:
    if pd.isna(x):
        return 0.0
    s = str(x).replace(',', '').strip()
    try:
        if s.endswith('万'):
            return float(s[:-1])
        if s.endswith('亿'):
            return float(s[:-1]) * 10000.0
        return float(s)
    except Exception:
        return 0.0


def parse_list_cell(x: Any) -> List[str]:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return []
    if isinstance(x, list):
        return [str(v).strip() for v in x if str(v).strip()]
    s = str(x).strip()
    if not s or s.lower() in {'nan','none','null'}:
        return []
    if s.startswith('[') and s.endswith(']'):
        try:
            import ast
            v = ast.literal_eval(s)
            if isinstance(v, list):
                return [str(t).strip() for t in v if str(t).strip()]
        except Exception:
            pass
    return [t.strip() for t in s.split('|') if t.strip()]


def mlb_transform_no_warn(mlb: MultiLabelBinarizer, seqs) -> np.ndarray:
    classes = list(getattr(mlb, "classes_", []))
    if not classes:
        return np.zeros((len(seqs), 0), dtype=float)
    idx = {c:i for i,c in enumerate(classes)}
    X = np.zeros((len(seqs), len(classes)), dtype=float)
    for r, labs in enumerate(seqs):
        if labs is None:
            continue
        for lab in labs:
            j = idx.get(lab)
            if j is not None:
                X[r, j] = 1.0
    return X

def safe_ln(x: float) -> float:
    try:
        x = float(x)
    except Exception:
        return np.nan
    return np.log(x) if x > 0 else np.nan


def metrics_on_ln_and_price(y_ln: np.ndarray, pred_ln: np.ndarray) -> Dict[str, float]:
    y_ln = np.asarray(y_ln, dtype=float)
    pred_ln = np.asarray(pred_ln, dtype=float)
    y = np.exp(y_ln)
    pred = np.exp(pred_ln)
    eps = 1e-12
    return {
        'R2': float(r2_score(y_ln, pred_ln)),
        'MAE': float(mean_absolute_error(y_ln, pred_ln)),
        'RMSE': float(np.sqrt(mean_squared_error(y_ln, pred_ln))),
        'MAPE': float(np.mean(np.abs((y - pred) / np.maximum(np.abs(y), eps)))),
        'SMAPE': float(np.mean(np.abs(y - pred) / np.maximum((np.abs(y) + np.abs(pred)) / 2.0, eps))),
    }


def spearman_rho(x: np.ndarray, y: np.ndarray) -> float:
    xr = pd.Series(x).rank(method='average').values
    yr = pd.Series(y).rank(method='average').values
    if np.std(xr) == 0 or np.std(yr) == 0:
        return np.nan
    return float(np.corrcoef(xr, yr)[0, 1])



def _prod_key(name: Any, supplier: Any) -> str:
    return f"PROD::{str(name).strip()}||SUP::{str(supplier).strip()}"


def _json_topk(post: Dict[str, float], topk: int = 10) -> str:
    if not post:
        return json.dumps({}, ensure_ascii=False)
    items = sorted(post.items(), key=lambda kv: kv[1], reverse=True)[:int(topk)]
    return json.dumps({str(k): float(v) for k, v in items}, ensure_ascii=False)


def _posterior_entropy(post: Dict[str, float]) -> float:
    if not post:
        return 0.0
    p = np.asarray(list(post.values()), dtype=float)
    p = p / (p.sum() + 1e-12)
    return float(-np.sum(p * np.log(p + 1e-12)))


def _expected_from_heat(post: Dict[str, float], heat_map: Dict[str, float]) -> float:
    if not post:
        return 0.0
    return float(sum(float(w) * float(heat_map.get(k, 0.0)) for k, w in post.items()))


def _renormalize_dict(d: Dict[str, float], topk: Optional[int] = None) -> Dict[str, float]:
    if not d:
        return {}
    items = [(str(k), float(v)) for k, v in d.items() if str(k).strip() and np.isfinite(float(v)) and float(v) > 0]
    if not items:
        return {}
    items = sorted(items, key=lambda kv: kv[1], reverse=True)
    if topk is not None:
        items = items[:int(topk)]
    s = float(sum(v for _, v in items))
    if s <= 1e-12:
        return {}
    return {k: float(v / s) for k, v in items}


def _avg_posteriors(posts: List[Dict[str, float]], topk: Optional[int] = None) -> Dict[str, float]:
    acc: Dict[str, float] = {}
    n = 0
    for post in posts:
        if not post:
            continue
        n += 1
        for k, v in post.items():
            acc[str(k)] = acc.get(str(k), 0.0) + float(v)
    if n <= 0:
        return {}
    for k in list(acc.keys()):
        acc[k] = acc[k] / float(n)
    return _renormalize_dict(acc, topk=topk)


def load_media_heat_maps_ln(media_path: str) -> Dict[str, Dict[str, float]]:
    """Read media heat and return ln(1+count)-scale maps, aligned with Appendix D/current program."""
    maps = {'total_log': {}, 'web_log': {}, 'news_log': {}, 'weixin_log': {}}
    if not media_path or not os.path.exists(media_path):
        return maps
    media = pd.read_excel(media_path).copy()
    media = media.rename(columns={'keyword': 'app'})
    if 'app' not in media.columns:
        return maps
    for c in ['sogou_web_results', 'sina_news_results', 'weixin_article_results']:
        if c not in media.columns:
            media[c] = 0.0
        media[c] = pd.to_numeric(media[c], errors='coerce').fillna(0.0)
    media['heat_web'] = np.log1p(media['sogou_web_results'])
    media['heat_news'] = np.log1p(media['sina_news_results'])
    media['heat_weixin'] = np.log1p(media['weixin_article_results'])
    media['heat_total'] = media['heat_web'] + media['heat_news'] + media['heat_weixin']
    for _, r in media.iterrows():
        app = str(r.get('app', '')).strip()
        if not app:
            continue
        maps['total_log'][app] = float(r.get('heat_total', 0.0))
        maps['web_log'][app] = float(r.get('heat_web', 0.0))
        maps['news_log'][app] = float(r.get('heat_news', 0.0))
        maps['weixin_log'][app] = float(r.get('heat_weixin', 0.0))
    return maps


def build_product_text_embedding(df_rows: pd.DataFrame, text_col: str = 'desc', dim: int = TEXT_EMB_DIM, min_df: int = TEXT_MIN_DF) -> pd.DataFrame:
    """Generate deterministic TF-IDF+SVD text embeddings used by the Product block."""
    d = df_rows.copy().reset_index(drop=True)
    if text_col not in d.columns:
        d[text_col] = ''
    d[text_col] = d[text_col].fillna('').astype(str)
    base = d[[c for c in ['name', 'supplier', 'prod_key'] if c in d.columns]].copy()
    try:
        tfidf = TfidfVectorizer(analyzer='char', ngram_range=(2, 4), min_df=int(min_df))
        X = tfidf.fit_transform(d[text_col].values)
        n_samples, n_features = X.shape
        if n_samples < 3 or n_features < 3:
            emb = np.zeros((len(d), 0), dtype=float)
        else:
            n_comp = int(min(dim, n_samples - 1, n_features - 1))
            n_comp = max(2, n_comp)
            svd = TruncatedSVD(n_components=n_comp, random_state=RANDOM_STATE)
            emb = svd.fit_transform(X).astype(float)
    except Exception:
        emb = np.zeros((len(d), 0), dtype=float)
    emb_df = pd.DataFrame(emb, columns=[f'textemb_{i:02d}' for i in range(emb.shape[1])])
    return pd.concat([base.reset_index(drop=True), emb_df], axis=1)


def knn_neighbors_exclude_self(X_train: np.ndarray, X_any: np.ndarray, k: int) -> Tuple[np.ndarray, np.ndarray]:
    """Cosine-KNN neighbors from X_train, excluding self when X_any is X_train."""
    if X_train.shape[0] < 2 or X_train.shape[1] == 0 or X_any.shape[0] == 0:
        return np.zeros((X_any.shape[0], 0), dtype=int), np.zeros((X_any.shape[0], 0), dtype=float)
    n_train = X_train.shape[0]
    k_use = int(min(max(1, k), n_train - 1))
    nn = NearestNeighbors(n_neighbors=min(k_use + 1, n_train), metric='cosine', algorithm='brute')
    nn.fit(X_train)
    dist, idx = nn.kneighbors(X_any, return_distance=True)
    sim = 1.0 - dist
    if X_any.shape[0] == n_train and np.allclose(X_any, X_train):
        idx2, sim2 = [], []
        for i in range(n_train):
            pairs = [(int(idx[i, j]), float(sim[i, j])) for j in range(idx.shape[1]) if int(idx[i, j]) != i]
            pairs = [(j, w) for j, w in pairs if w >= SIM_MIN_COS][:k_use]
            idx2.append([j for j, _ in pairs])
            sim2.append([w for _, w in pairs])
        maxk = max([len(x) for x in idx2]) if idx2 else 0
        idx_arr = np.zeros((n_train, maxk), dtype=int)
        sim_arr = np.zeros((n_train, maxk), dtype=float)
        for i in range(n_train):
            for j in range(len(idx2[i])):
                idx_arr[i, j] = idx2[i][j]
                sim_arr[i, j] = sim2[i][j]
        return idx_arr, sim_arr
    idx = idx[:, :k_use]
    sim = sim[:, :k_use]
    sim = np.where(sim >= SIM_MIN_COS, sim, 0.0)
    return idx.astype(int), sim.astype(float)


def neighbor_label_posterior(neighbor_label_lists: List[List[str]], neighbor_weights: np.ndarray) -> Dict[str, float]:
    """Build posterior over KG labels from text-similarity competitors; neighbor weight is fractionally assigned over labels."""
    post: Dict[str, float] = {}
    for labs, w in zip(neighbor_label_lists, neighbor_weights):
        labs = [str(x).strip() for x in (labs or []) if str(x).strip()]
        if not labs:
            continue
        ww = float(max(w, 0.0)) / float(len(labs))
        for lab in labs:
            post[lab] = post.get(lab, 0.0) + ww
    return _renormalize_dict(post)


def step0_write_competitor_outputs_from_desc(df_rows: pd.DataFrame, heat_maps: Dict[str, Dict[str, float]], out_dir: str, k: int = SIM_KNN_K, min_cos: float = SIM_MIN_COS, topk_labels: int = METAPATH_TOPK) -> pd.DataFrame:
    """Generate STEP0_product_competitor_summary.csv from text-similarity neighbors and KG app/src labels.

    This restores the upstream competitor-neighborhood reasoning process that the current manuscript tables
    already consume through comp_expected_heat_total, comp_app_entropy, comp_src_entropy, and top-1 variables.
    """
    ensure_dir(out_dir)
    d = df_rows.copy().reset_index(drop=True)
    d['name'] = d['name'].astype(str)
    d['supplier'] = d['supplier'].astype(str)
    if 'desc' not in d.columns:
        d['desc'] = ''
    d['desc'] = d['desc'].fillna('').astype(str)
    if 'app_list' not in d.columns:
        d['app_list'] = [[] for _ in range(len(d))]
    if 'src_list' not in d.columns:
        d['src_list'] = [[] for _ in range(len(d))]
    d['app_list'] = d['app_list'].apply(lambda x: x if isinstance(x, list) else parse_list_cell(x))
    d['src_list'] = d['src_list'].apply(lambda x: x if isinstance(x, list) else parse_list_cell(x))
    if 'prod_key' not in d.columns:
        d['prod_key'] = d.apply(lambda r: _prod_key(r['name'], r['supplier']), axis=1)

    emb = build_product_text_embedding(d, text_col='desc', dim=TEXT_EMB_DIM, min_df=TEXT_MIN_DF)
    emb_path = os.path.join(out_dir, 'STEP0_product_textemb.csv')
    emb.to_csv(emb_path, index=False, encoding='utf-8-sig')
    emb_cols = [c for c in emb.columns if c.startswith('textemb_')]
    X = emb[emb_cols].astype(float).values if emb_cols else np.zeros((len(d), 0), dtype=float)
    nbr_idx, nbr_sim = knn_neighbors_exclude_self(X, X, k=k)

    app_lists = [parse_list_cell(x) for x in d['app_list'].tolist()]
    src_lists = [parse_list_cell(x) for x in d['src_list'].tolist()]
    prod_keys = [_prod_key(r['name'], r['supplier']) for _, r in d.iterrows()]
    rows, missing_app_rows, missing_src_rows, edge_rows = [], [], [], []
    for i in range(len(d)):
        idx = nbr_idx[i] if nbr_idx.size else np.asarray([], dtype=int)
        sim = nbr_sim[i] if nbr_sim.size else np.asarray([], dtype=float)
        valid = [(int(j), float(w)) for j, w in zip(idx.tolist(), sim.tolist()) if float(w) > 0]
        neigh_apps = [app_lists[j] for j, _ in valid]
        neigh_srcs = [src_lists[j] for j, _ in valid]
        weights = np.asarray([w for _, w in valid], dtype=float)
        post_app = neighbor_label_posterior(neigh_apps, weights)
        post_src = neighbor_label_posterior(neigh_srcs, weights)
        comp = {prod_keys[j]: float(w) for j, w in valid}
        for j, w in valid:
            edge_rows.append({'u': prod_keys[i], 'v': prod_keys[j], 'w': float(w)})
        obs_app, obs_src = set(app_lists[i]), set(src_lists[i])
        inferred_app = [k0 for k0, _ in sorted(post_app.items(), key=lambda kv: kv[1], reverse=True)[:topk_labels]]
        inferred_src = [k0 for k0, _ in sorted(post_src.items(), key=lambda kv: kv[1], reverse=True)[:topk_labels]]
        rows.append({
            'name': str(d.iloc[i]['name']),
            'supplier': str(d.iloc[i]['supplier']),
            'prod_key': prod_keys[i],
            'comp_app_entropy': _posterior_entropy(post_app),
            'comp_src_entropy': _posterior_entropy(post_src),
            'comp_app_top1_prob': max(post_app.values()) if post_app else 0.0,
            'comp_src_top1_prob': max(post_src.values()) if post_src else 0.0,
            'comp_expected_heat_total': _expected_from_heat(post_app, heat_maps.get('total_log', {})),
            'comp_top_products_json': _json_topk(comp, topk=k),
            'comp_inferred_apps_json': _json_topk(post_app, topk=topk_labels),
            'comp_inferred_srcs_json': _json_topk(post_src, topk=topk_labels),
        })
        missing_app_rows.append({'name': str(d.iloc[i]['name']), 'supplier': str(d.iloc[i]['supplier']), 'observed_count': int(len(obs_app)), 'inferred_topk': json.dumps(inferred_app, ensure_ascii=False), 'missing_topk': json.dumps([x for x in inferred_app if x not in obs_app], ensure_ascii=False)})
        missing_src_rows.append({'name': str(d.iloc[i]['name']), 'supplier': str(d.iloc[i]['supplier']), 'observed_count': int(len(obs_src)), 'inferred_topk': json.dumps(inferred_src, ensure_ascii=False), 'missing_topk': json.dumps([x for x in inferred_src if x not in obs_src], ensure_ascii=False)})

    df_sum = pd.DataFrame(rows)
    df_sum.to_csv(os.path.join(out_dir, 'STEP0_product_competitor_summary.csv'), index=False, encoding='utf-8-sig')
    pd.DataFrame(edge_rows).to_csv(os.path.join(out_dir, 'STEP0_product_similarity_graph_edges.csv'), index=False, encoding='utf-8-sig')
    pd.DataFrame(missing_app_rows).to_csv(os.path.join(out_dir, 'STEP0_completion_product_missing_apps.csv'), index=False, encoding='utf-8-sig')
    pd.DataFrame(missing_src_rows).to_csv(os.path.join(out_dir, 'STEP0_completion_product_missing_srcs.csv'), index=False, encoding='utf-8-sig')
    keep = ['name','supplier','comp_app_entropy','comp_src_entropy','comp_app_top1_prob','comp_src_top1_prob','comp_expected_heat_total']
    return df_sum[keep].copy()


def ensure_product_reasoning_input_files(df_rows: pd.DataFrame, data_dir: str, media_path: str, out_dir: Optional[str] = None, force: bool = False) -> None:
    """Ensure current program can reproduce the precomputed Product-block inputs.

    Existing files are preserved by default. If missing, the function regenerates them.
    """
    ensure_dir(data_dir)
    text_path = os.path.join(data_dir, 'STEP0_product_textemb.csv')
    comp_path = os.path.join(data_dir, 'STEP0_product_competitor_summary.csv')
    heat_maps = load_media_heat_maps_ln(media_path)
    if force or not os.path.exists(text_path) or not os.path.exists(comp_path):
        print('[KG reasoning] generating missing STEP0 product text embeddings / competitor summaries', flush=True)
        step0_write_competitor_outputs_from_desc(df_rows, heat_maps=heat_maps, out_dir=data_dir, k=SIM_KNN_K, min_cos=SIM_MIN_COS, topk_labels=METAPATH_TOPK)
    if out_dir:
        ensure_dir(out_dir)
        # keep an auditable copy in result directory without changing the data_dir source files
        for fn in ['STEP0_product_textemb.csv', 'STEP0_product_competitor_summary.csv', 'STEP0_product_similarity_graph_edges.csv', 'STEP0_completion_product_missing_apps.csv', 'STEP0_completion_product_missing_srcs.csv']:
            src = os.path.join(data_dir, fn)
            dst = os.path.join(out_dir, fn)
            if os.path.exists(src) and (force or not os.path.exists(dst)):
                try:
                    pd.read_csv(src).to_csv(dst, index=False, encoding='utf-8-sig')
                except Exception:
                    pass


class KGPPRReasoner:
    """Heterogeneous KG Personalized PageRank on Supplier–Product–Source/Application graph."""
    def __init__(self, df_rows: pd.DataFrame, app_heat_total: Dict[str, float], topk_apps: int = PPR_TOPK_APPS, topk_srcs: int = PPR_TOPK_SRCS):
        if nx is None:
            raise RuntimeError('networkx is not installed; cannot run PPR reasoning')
        self.app_heat_total = app_heat_total or {}
        self.topk_apps = int(topk_apps)
        self.topk_srcs = int(topk_srcs)
        self.G = nx.DiGraph()
        self._cache: Dict[Tuple[str, float, str], Dict[str, float]] = {}
        self._build_graph(df_rows)
        self._prepare_heat_z()

    def _build_graph(self, df_rows: pd.DataFrame) -> None:
        d = df_rows.copy()
        for _, r in d.iterrows():
            s = str(r.get('supplier', '')).strip()
            name = str(r.get('name', '')).strip()
            if not s or not name:
                continue
            sup = f'SUP::{s}'
            prod = _prod_key(name, s)
            self.G.add_edge(sup, prod, weight=1.0, etype='provide_data')
            self.G.add_edge(prod, sup, weight=1.0, etype='provide_data')
            for a in parse_list_cell(r.get('app_list', [])):
                app = f'APP::{a}'
                self.G.add_edge(prod, app, weight=1.0, etype='applied_to')
                self.G.add_edge(app, prod, weight=1.0, etype='applied_to')
            for src in parse_list_cell(r.get('src_list', [])):
                srcn = f'SRC::{src}'
                self.G.add_edge(srcn, prod, weight=1.0, etype='source_industry')
                self.G.add_edge(prod, srcn, weight=1.0, etype='source_industry')

    def _prepare_heat_z(self) -> None:
        apps = [n.replace('APP::', '') for n in self.G.nodes() if str(n).startswith('APP::')]
        vals = np.asarray([self.app_heat_total.get(a, 0.0) for a in apps], dtype=float)
        if len(vals) == 0 or np.std(vals) < 1e-12:
            self.heat_z = {a: 0.0 for a in apps}
            return
        z = (vals - vals.mean()) / (vals.std() + 1e-12)
        z = np.clip(z, -2.0, 2.0)
        self.heat_z = {apps[i]: float(z[i]) for i in range(len(apps))}

    def _ppr(self, start_node: str, alpha: float, personalization: Optional[Dict[str, float]] = None, tag: str = 'struct') -> Dict[str, float]:
        if start_node not in self.G:
            return {}
        key = (start_node, float(alpha), tag)
        if key in self._cache:
            return self._cache[key]
        if personalization is None:
            personalization = {n: 0.0 for n in self.G.nodes()}
            personalization[start_node] = 1.0
        else:
            pers = {n: 0.0 for n in self.G.nodes()}
            for k, v in personalization.items():
                if k in pers:
                    pers[k] = float(v)
            personalization = pers
        pr = nx.pagerank(self.G, alpha=alpha, personalization=personalization, weight='weight', max_iter=MAX_NX_PAGERANK_ITERS, tol=NX_PAGERANK_TOL)
        out = {str(k): float(v) for k, v in pr.items()}
        self._cache[key] = out
        return out

    def _extract(self, pr: Dict[str, float], prefix: str, topk: int) -> Dict[str, float]:
        raw = {str(n).replace(prefix, ''): float(v) for n, v in pr.items() if str(n).startswith(prefix)}
        return _renormalize_dict(raw, topk=topk)

    def _heat_personalization(self, start_node: str, prior_strength: float, teleport_mass: float) -> Dict[str, float]:
        apps = [n.replace('APP::', '') for n in self.G.nodes() if str(n).startswith('APP::')]
        pers = {start_node: 1.0 - float(teleport_mass)}
        if apps and teleport_mass > 0:
            w = np.asarray([np.exp(float(prior_strength) * float(self.heat_z.get(a, 0.0))) for a in apps], dtype=float)
            w = w / (w.sum() + 1e-12)
            for a, p in zip(apps, w):
                pers[f'APP::{a}'] = pers.get(f'APP::{a}', 0.0) + float(teleport_mass) * float(p)
        return pers

    def infer_product_apps(self, name: str, supplier: str, alpha: float = PPR_ALPHA, prior_strength: float = PPR_PRIOR_STRENGTH) -> Dict[str, float]:
        node = _prod_key(name, supplier)
        pr = self._ppr(node, alpha=alpha)
        raw = self._extract(pr, 'APP::', self.topk_apps)
        if not raw:
            return {}
        post = {a: float(v) * float(np.exp(float(prior_strength) * float(self.heat_z.get(a, 0.0)))) for a, v in raw.items()}
        return _renormalize_dict(post, topk=self.topk_apps)

    def infer_product_apps_structural(self, name: str, supplier: str, alpha: float = PPR_ALPHA) -> Dict[str, float]:
        return self._extract(self._ppr(_prod_key(name, supplier), alpha=alpha), 'APP::', self.topk_apps)

    def infer_product_apps_heat_teleport(self, name: str, supplier: str, alpha: float = PPR_ALPHA, prior_strength: float = PPR_PRIOR_STRENGTH, teleport_mass: float = PPR_HEAT_TELEPORT_MASS) -> Dict[str, float]:
        node = _prod_key(name, supplier)
        pers = self._heat_personalization(node, prior_strength=prior_strength, teleport_mass=teleport_mass)
        return self._extract(self._ppr(node, alpha=alpha, personalization=pers, tag=f'heattele_{teleport_mass:.3f}'), 'APP::', self.topk_apps)

    def infer_product_srcs(self, name: str, supplier: str, alpha: float = PPR_ALPHA) -> Dict[str, float]:
        return self._extract(self._ppr(_prod_key(name, supplier), alpha=alpha), 'SRC::', self.topk_srcs)

    def infer_supplier_apps_structural(self, supplier: str, alpha: float = PPR_ALPHA) -> Dict[str, float]:
        return self._extract(self._ppr(f'SUP::{supplier}', alpha=alpha), 'APP::', self.topk_apps)

    def infer_supplier_apps_heat_teleport(self, supplier: str, alpha: float = PPR_ALPHA, prior_strength: float = PPR_PRIOR_STRENGTH, teleport_mass: float = PPR_HEAT_TELEPORT_MASS) -> Dict[str, float]:
        node = f'SUP::{supplier}'
        pers = self._heat_personalization(node, prior_strength=prior_strength, teleport_mass=teleport_mass)
        return self._extract(self._ppr(node, alpha=alpha, personalization=pers, tag=f'sup_heattele_{teleport_mass:.3f}'), 'APP::', self.topk_apps)

    def infer_src_apps_structural(self, src: str, alpha: float = PPR_ALPHA) -> Dict[str, float]:
        return self._extract(self._ppr(f'SRC::{src}', alpha=alpha), 'APP::', self.topk_apps)

    def infer_src_apps_heat_teleport(self, src: str, alpha: float = PPR_ALPHA, prior_strength: float = PPR_PRIOR_STRENGTH, teleport_mass: float = PPR_HEAT_TELEPORT_MASS) -> Dict[str, float]:
        node = f'SRC::{src}'
        pers = self._heat_personalization(node, prior_strength=prior_strength, teleport_mass=teleport_mass)
        return self._extract(self._ppr(node, alpha=alpha, personalization=pers, tag=f'src_heattele_{teleport_mass:.3f}'), 'APP::', self.topk_apps)

    def infer_supplier_apps(self, supplier: str, alpha: float = PPR_ALPHA) -> Dict[str, float]:
        return self.infer_supplier_apps_structural(supplier, alpha=alpha)

    def infer_supplier_srcs(self, supplier: str, alpha: float = PPR_ALPHA) -> Dict[str, float]:
        return self._extract(self._ppr(f'SUP::{supplier}', alpha=alpha), 'SRC::', self.topk_srcs)

    def infer_src_apps(self, src: str, alpha: float = PPR_ALPHA) -> Dict[str, float]:
        return self.infer_src_apps_structural(src, alpha=alpha)

    def infer_app_srcs(self, app: str, alpha: float = PPR_ALPHA) -> Dict[str, float]:
        return self._extract(self._ppr(f'APP::{app}', alpha=alpha), 'SRC::', self.topk_srcs)

    def expected_heat(self, post_apps: Dict[str, float]) -> float:
        return _expected_from_heat(post_apps, self.app_heat_total)


class MetaPathPPRReasoner:
    """Meta-path constrained PPR on SUP-APP, SUP-SRC, and SRC-APP bipartite graphs."""
    def __init__(self, df_rows: pd.DataFrame, app_heat_total: Dict[str, float]):
        if nx is None:
            raise RuntimeError('networkx is not installed; cannot run MPPR reasoning')
        self.app_heat_total = app_heat_total or {}
        self.G_sa, self.G_ss, self.G_ca = self._build_graphs(df_rows)
        self._cache: Dict[Tuple[int, str, float], Dict[str, float]] = {}
        self._prepare_heat_z()

    def _build_graphs(self, df_rows: pd.DataFrame):
        G_sa, G_ss, G_ca = nx.Graph(), nx.Graph(), nx.Graph()
        for _, r in df_rows.iterrows():
            sup = str(r.get('supplier', '')).strip()
            if not sup:
                continue
            apps = list(dict.fromkeys(parse_list_cell(r.get('app_list', []))))
            srcs = list(dict.fromkeys(parse_list_cell(r.get('src_list', []))))
            for a in apps:
                self._add_weight(G_sa, f'SUP::{sup}', f'APP::{a}', 1.0)
            for src in srcs:
                self._add_weight(G_ss, f'SUP::{sup}', f'SRC::{src}', 1.0)
            for src in srcs:
                for a in apps:
                    self._add_weight(G_ca, f'SRC::{src}', f'APP::{a}', 1.0)
        return G_sa, G_ss, G_ca

    @staticmethod
    def _add_weight(G, u: str, v: str, w: float) -> None:
        if G.has_edge(u, v):
            G[u][v]['weight'] += float(w)
        else:
            G.add_edge(u, v, weight=float(w))

    def _prepare_heat_z(self) -> None:
        apps = sorted({n.replace('APP::', '') for G in [self.G_sa, self.G_ca] for n in G.nodes() if str(n).startswith('APP::')})
        vals = np.asarray([self.app_heat_total.get(a, 0.0) for a in apps], dtype=float)
        if len(vals) == 0 or np.std(vals) < 1e-12:
            self.heat_z = {a: 0.0 for a in apps}
        else:
            z = (vals - vals.mean()) / (vals.std() + 1e-12)
            z = np.clip(z, -2.0, 2.0)
            self.heat_z = {apps[i]: float(z[i]) for i in range(len(apps))}

    def _ppr(self, G, start_node: str, alpha: float = METAPATH_ALPHA, personalization: Optional[Dict[str, float]] = None) -> Dict[str, float]:
        if start_node not in G:
            return {}
        key = (id(G), start_node, float(alpha)) if personalization is None else (id(G), start_node + '__custom__' + str(hash(tuple(sorted(personalization.items())))), float(alpha))
        if key in self._cache:
            return self._cache[key]
        if personalization is None:
            personalization = {n: 0.0 for n in G.nodes()}
            personalization[start_node] = 1.0
        else:
            pers = {n: 0.0 for n in G.nodes()}
            for k, v in personalization.items():
                if k in pers:
                    pers[k] = float(v)
            personalization = pers
        pr = nx.pagerank(G, alpha=alpha, personalization=personalization, weight='weight', max_iter=MAX_NX_PAGERANK_ITERS, tol=NX_PAGERANK_TOL)
        out = {str(k): float(v) for k, v in pr.items()}
        self._cache[key] = out
        return out

    def _extract(self, pr: Dict[str, float], prefix: str, topk: int) -> Dict[str, float]:
        return _renormalize_dict({n.replace(prefix, ''): v for n, v in pr.items() if str(n).startswith(prefix)}, topk=topk)

    def _posterior_with_heat_prior(self, raw_app: Dict[str, float], prior_strength: float) -> Dict[str, float]:
        post = {a: float(sc) * float(np.exp(float(prior_strength) * float(self.heat_z.get(a, 0.0)))) for a, sc in raw_app.items()}
        return _renormalize_dict(post, topk=METAPATH_TOPK)

    def _heat_personalization(self, G, start_node: str, prior_strength: float, teleport_mass: float) -> Dict[str, float]:
        apps = [n.replace('APP::', '') for n in G.nodes() if str(n).startswith('APP::')]
        pers = {start_node: 1.0 - float(teleport_mass)}
        if apps and teleport_mass > 0:
            w = np.asarray([np.exp(float(prior_strength) * float(self.heat_z.get(a, 0.0))) for a in apps], dtype=float)
            w = w / (w.sum() + 1e-12)
            for a, p in zip(apps, w):
                pers[f'APP::{a}'] = pers.get(f'APP::{a}', 0.0) + float(teleport_mass) * float(p)
        return pers

    def sup_to_apps_structural(self, supplier: str, alpha: float = METAPATH_ALPHA, topk: int = METAPATH_TOPK) -> Dict[str, float]:
        return self._extract(self._ppr(self.G_sa, f'SUP::{supplier}', alpha=alpha), 'APP::', topk)

    def src_to_apps_structural(self, src: str, alpha: float = METAPATH_ALPHA, topk: int = METAPATH_TOPK) -> Dict[str, float]:
        return self._extract(self._ppr(self.G_ca, f'SRC::{src}', alpha=alpha), 'APP::', topk)

    def sup_to_apps_heat_teleport(self, supplier: str, alpha: float = METAPATH_ALPHA, topk: int = METAPATH_TOPK, prior_strength: float = METAPATH_PRIOR_STRENGTH_APPHEAT, teleport_mass: float = METAPATH_HEAT_TELEPORT_MASS) -> Dict[str, float]:
        node = f'SUP::{supplier}'
        pers = self._heat_personalization(self.G_sa, node, prior_strength=prior_strength, teleport_mass=teleport_mass)
        return self._extract(self._ppr(self.G_sa, node, alpha=alpha, personalization=pers), 'APP::', topk)

    def src_to_apps_heat_teleport(self, src: str, alpha: float = METAPATH_ALPHA, topk: int = METAPATH_TOPK, prior_strength: float = METAPATH_PRIOR_STRENGTH_APPHEAT, teleport_mass: float = METAPATH_HEAT_TELEPORT_MASS) -> Dict[str, float]:
        node = f'SRC::{src}'
        pers = self._heat_personalization(self.G_ca, node, prior_strength=prior_strength, teleport_mass=teleport_mass)
        return self._extract(self._ppr(self.G_ca, node, alpha=alpha, personalization=pers), 'APP::', topk)

    def sup_to_apps(self, supplier: str, alpha: float = METAPATH_ALPHA, topk: int = METAPATH_TOPK, prior_strength: float = METAPATH_PRIOR_STRENGTH_APPHEAT) -> Dict[str, float]:
        raw = self._extract(self._ppr(self.G_sa, f'SUP::{supplier}', alpha=alpha), 'APP::', topk)
        return self._posterior_with_heat_prior(raw, prior_strength)

    def sup_to_srcs(self, supplier: str, alpha: float = METAPATH_ALPHA, topk: int = METAPATH_TOPK) -> Dict[str, float]:
        return self._extract(self._ppr(self.G_ss, f'SUP::{supplier}', alpha=alpha), 'SRC::', topk)

    def src_to_apps(self, src: str, alpha: float = METAPATH_ALPHA, topk: int = METAPATH_TOPK, prior_strength: float = METAPATH_PRIOR_STRENGTH_APPHEAT) -> Dict[str, float]:
        raw = self._extract(self._ppr(self.G_ca, f'SRC::{src}', alpha=alpha), 'APP::', topk)
        return self._posterior_with_heat_prior(raw, prior_strength)

    def app_to_srcs(self, app: str, alpha: float = METAPATH_ALPHA, topk: int = METAPATH_TOPK) -> Dict[str, float]:
        return self._extract(self._ppr(self.G_ca, f'APP::{app}', alpha=alpha), 'SRC::', topk)

    def app_to_sups(self, app: str, alpha: float = METAPATH_ALPHA, topk: int = METAPATH_TOPK) -> Dict[str, float]:
        return self._extract(self._ppr(self.G_sa, f'APP::{app}', alpha=alpha), 'SUP::', topk)

    def expected_app_heat(self, post_apps: Dict[str, float]) -> float:
        return _expected_from_heat(post_apps, self.app_heat_total)


def write_kg_reasoning_outputs(df: pd.DataFrame, out_dir: str, media_path: str, heat_ablation: bool = True, ppr_heat_teleport_mass: float = PPR_HEAT_TELEPORT_MASS, metapath_heat_teleport_mass: float = METAPATH_HEAT_TELEPORT_MASS) -> None:
    """Run and export auditable KG reasoning outputs without changing model/table behavior."""
    ensure_dir(out_dir)
    if nx is None:
        write_text(os.path.join(out_dir, 'KG_REASONING_SKIPPED.txt'), 'networkx is unavailable; KG PPR/MPPR outputs were not generated.\n')
        return
    d = df.copy()
    if 'prod_key' not in d.columns:
        d['prod_key'] = d.apply(lambda r: _prod_key(r.get('name', ''), r.get('supplier', '')), axis=1)
    d['app_list'] = d['app_list'].apply(lambda x: x if isinstance(x, list) else parse_list_cell(x))
    d['src_list'] = d['src_list'].apply(lambda x: x if isinstance(x, list) else parse_list_cell(x))
    heat_maps = load_media_heat_maps_ln(media_path)
    app_heat = heat_maps.get('total_log', {})
    kg = KGPPRReasoner(d, app_heat_total=app_heat)
    mp = MetaPathPPRReasoner(d, app_heat_total=app_heat)

    # Product-level heterogeneous-graph PPR outputs
    prod_rows = []
    for _, r in d.drop_duplicates('prod_key').iterrows():
        name, sup, pk = str(r['name']), str(r['supplier']), str(r['prod_key'])
        post_app = kg.infer_product_apps(name, sup)
        post_src = kg.infer_product_srcs(name, sup)
        post_app_struct = kg.infer_product_apps_structural(name, sup) if heat_ablation else post_app
        post_app_tele = kg.infer_product_apps_heat_teleport(name, sup, teleport_mass=ppr_heat_teleport_mass) if heat_ablation else post_app
        prod_rows.append({
            'prod_key': pk,
            'prod_ppr_expected_heat_total': kg.expected_heat(post_app),
            'prod_ppr_expected_heat_total_struct': kg.expected_heat(post_app_struct),
            'prod_ppr_expected_heat_total_tele': kg.expected_heat(post_app_tele),
            'prod_ppr_expheat_delta_tele_minus_struct': kg.expected_heat(post_app_tele) - kg.expected_heat(post_app_struct),
            'prod_ppr_app_entropy': _posterior_entropy(post_app),
            'prod_ppr_src_entropy': _posterior_entropy(post_src),
            'prod_ppr_app_top1_prob': max(post_app.values()) if post_app else 0.0,
            'prod_ppr_src_top1_prob': max(post_src.values()) if post_src else 0.0,
            'prod_ppr_top_apps_json': _json_topk(post_app, PPR_TOPK_APPS),
            'prod_ppr_top_srcs_json': _json_topk(post_src, PPR_TOPK_SRCS),
        })
    df_prod_ppr = pd.DataFrame(prod_rows)
    df_prod_ppr.to_csv(os.path.join(out_dir, 'STEP2_KG_reasoning_product_ppr.csv'), index=False, encoding='utf-8-sig')

    # Product-level meta-path constrained PPR outputs
    mp_rows = []
    for _, r in d.drop_duplicates('prod_key').iterrows():
        sup = str(r['supplier'])
        srcs = parse_list_cell(r.get('src_list', []))
        apps = parse_list_cell(r.get('app_list', []))
        post_app = _avg_posteriors([mp.sup_to_apps(sup)] + [mp.src_to_apps(src) for src in srcs], topk=METAPATH_TOPK)
        post_app_struct = _avg_posteriors([mp.sup_to_apps_structural(sup)] + [mp.src_to_apps_structural(src) for src in srcs], topk=METAPATH_TOPK)
        post_app_tele = _avg_posteriors([mp.sup_to_apps_heat_teleport(sup, teleport_mass=metapath_heat_teleport_mass)] + [mp.src_to_apps_heat_teleport(src, teleport_mass=metapath_heat_teleport_mass) for src in srcs], topk=METAPATH_TOPK) if heat_ablation else post_app_struct
        post_src = _avg_posteriors([mp.sup_to_srcs(sup)] + [mp.app_to_srcs(app) for app in apps], topk=METAPATH_TOPK)
        mp_rows.append({
            'prod_key': str(r['prod_key']),
            'prod_mp_ppr_expected_heat_total': mp.expected_app_heat(post_app),
            'prod_mp_ppr_expected_heat_total_struct': mp.expected_app_heat(post_app_struct),
            'prod_mp_ppr_expected_heat_total_tele': mp.expected_app_heat(post_app_tele),
            'prod_mp_ppr_expheat_delta_tele_minus_struct': mp.expected_app_heat(post_app_tele) - mp.expected_app_heat(post_app_struct),
            'prod_mp_ppr_app_entropy': _posterior_entropy(post_app),
            'prod_mp_ppr_src_entropy': _posterior_entropy(post_src),
            'prod_mp_ppr_app_top1_prob': max(post_app.values()) if post_app else 0.0,
            'prod_mp_ppr_src_top1_prob': max(post_src.values()) if post_src else 0.0,
            'prod_mp_ppr_top_apps_json': _json_topk(post_app, METAPATH_TOPK),
            'prod_mp_ppr_top_srcs_json': _json_topk(post_src, METAPATH_TOPK),
        })
    df_prod_mp = pd.DataFrame(mp_rows)
    df_prod_mp.to_csv(os.path.join(out_dir, 'STEP2_KG_reasoning_product_mppr.csv'), index=False, encoding='utf-8-sig')

    # Entity-level outputs for audit and Appendix E support
    suppliers = sorted(d['supplier'].astype(str).unique())
    sup_rows = []
    for sup in suppliers:
        post_app = kg.infer_supplier_apps(sup)
        post_src = kg.infer_supplier_srcs(sup)
        mp_app = mp.sup_to_apps(sup)
        mp_src = mp.sup_to_srcs(sup)
        post_app_struct = kg.infer_supplier_apps_structural(sup)
        post_app_tele = kg.infer_supplier_apps_heat_teleport(sup, teleport_mass=ppr_heat_teleport_mass) if heat_ablation else post_app_struct
        mp_app_struct = mp.sup_to_apps_structural(sup)
        mp_app_tele = mp.sup_to_apps_heat_teleport(sup, teleport_mass=metapath_heat_teleport_mass) if heat_ablation else mp_app_struct
        sup_rows.append({'supplier': sup, 'ppr_expected_heat_total': kg.expected_heat(post_app), 'ppr_expected_heat_total_struct': kg.expected_heat(post_app_struct), 'ppr_expected_heat_total_tele': kg.expected_heat(post_app_tele), 'ppr_expheat_delta_tele_minus_struct': kg.expected_heat(post_app_tele) - kg.expected_heat(post_app_struct), 'ppr_app_entropy': _posterior_entropy(post_app), 'ppr_src_entropy': _posterior_entropy(post_src), 'ppr_app_top1_prob': max(post_app.values()) if post_app else 0.0, 'ppr_src_top1_prob': max(post_src.values()) if post_src else 0.0, 'mp_ppr_expected_heat_total': mp.expected_app_heat(mp_app), 'mp_ppr_expected_heat_total_struct': mp.expected_app_heat(mp_app_struct), 'mp_ppr_expected_heat_total_tele': mp.expected_app_heat(mp_app_tele), 'mp_ppr_expheat_delta_tele_minus_struct': mp.expected_app_heat(mp_app_tele) - mp.expected_app_heat(mp_app_struct), 'mp_ppr_app_entropy': _posterior_entropy(mp_app), 'mp_ppr_src_entropy': _posterior_entropy(mp_src), 'mp_ppr_app_top1_prob': max(mp_app.values()) if mp_app else 0.0, 'mp_ppr_src_top1_prob': max(mp_src.values()) if mp_src else 0.0, 'ppr_top_apps_json': _json_topk(post_app), 'ppr_top_srcs_json': _json_topk(post_src), 'mp_ppr_top_apps_json': _json_topk(mp_app), 'mp_ppr_top_srcs_json': _json_topk(mp_src)})
    pd.DataFrame(sup_rows).to_csv(os.path.join(out_dir, 'STEP2_mechanisms_supplier.csv'), index=False, encoding='utf-8-sig')

    srcs_all = sorted({x for labs in d['src_list'] for x in parse_list_cell(labs)})
    src_rows = []
    for src in srcs_all:
        post_app = kg.infer_src_apps(src)
        mp_app = mp.src_to_apps(src)
        post_app_struct = kg.infer_src_apps_structural(src)
        post_app_tele = kg.infer_src_apps_heat_teleport(src, teleport_mass=ppr_heat_teleport_mass) if heat_ablation else post_app_struct
        mp_app_struct = mp.src_to_apps_structural(src)
        mp_app_tele = mp.src_to_apps_heat_teleport(src, teleport_mass=metapath_heat_teleport_mass) if heat_ablation else mp_app_struct
        src_rows.append({'src': src, 'ppr_expected_heat_total': kg.expected_heat(post_app), 'ppr_expected_heat_total_struct': kg.expected_heat(post_app_struct), 'ppr_expected_heat_total_tele': kg.expected_heat(post_app_tele), 'ppr_expheat_delta_tele_minus_struct': kg.expected_heat(post_app_tele) - kg.expected_heat(post_app_struct), 'ppr_app_entropy': _posterior_entropy(post_app), 'ppr_app_top1_prob': max(post_app.values()) if post_app else 0.0, 'mp_ppr_expected_heat_total': mp.expected_app_heat(mp_app), 'mp_ppr_expected_heat_total_struct': mp.expected_app_heat(mp_app_struct), 'mp_ppr_expected_heat_total_tele': mp.expected_app_heat(mp_app_tele), 'mp_ppr_expheat_delta_tele_minus_struct': mp.expected_app_heat(mp_app_tele) - mp.expected_app_heat(mp_app_struct), 'mp_ppr_app_entropy': _posterior_entropy(mp_app), 'mp_ppr_app_top1_prob': max(mp_app.values()) if mp_app else 0.0, 'ppr_top_apps_json': _json_topk(post_app), 'mp_ppr_top_apps_json': _json_topk(mp_app)})
    pd.DataFrame(src_rows).to_csv(os.path.join(out_dir, 'STEP2_mechanisms_src.csv'), index=False, encoding='utf-8-sig')

    apps_all = sorted({x for labs in d['app_list'] for x in parse_list_cell(labs)})
    app_rows = []
    for app in apps_all:
        post_src = kg.infer_app_srcs(app)
        mp_src = mp.app_to_srcs(app)
        mp_sup = mp.app_to_sups(app)
        app_rows.append({'app': app, 'heat_total': float(app_heat.get(app, 0.0)), 'ppr_src_entropy': _posterior_entropy(post_src), 'ppr_src_top1_prob': max(post_src.values()) if post_src else 0.0, 'mp_ppr_src_entropy': _posterior_entropy(mp_src), 'mp_ppr_sup_entropy': _posterior_entropy(mp_sup), 'mp_ppr_src_top1_prob': max(mp_src.values()) if mp_src else 0.0, 'mp_ppr_sup_top1_prob': max(mp_sup.values()) if mp_sup else 0.0, 'ppr_top_srcs_json': _json_topk(post_src), 'mp_ppr_top_srcs_json': _json_topk(mp_src), 'mp_ppr_top_suppliers_json': _json_topk(mp_sup)})
    pd.DataFrame(app_rows).to_csv(os.path.join(out_dir, 'STEP2_mechanisms_app.csv'), index=False, encoding='utf-8-sig')

    # Combined dataset for downstream inspection/plots. This does not feed back into Tables 1-7.
    mech = d.merge(df_prod_ppr, on='prod_key', how='left').merge(df_prod_mp, on='prod_key', how='left')
    mech.to_csv(os.path.join(out_dir, 'STEP2_dataset_with_mechanisms.csv'), index=False, encoding='utf-8-sig')
    write_text(os.path.join(out_dir, 'KG_REASONING_README.txt'), 'Generated KG reasoning outputs: product PPR, product MPPR, supplier/src/app mechanisms, and STEP2_dataset_with_mechanisms.csv. Tables 1-6 keep the existing Bayesian Ridge and KG-derived competitor-neighborhood inputs. Table 7 now reports PPR/MPPR structural-versus-heat-teleport robustness. Appendix E tables report PPR/MPPR variable definitions, mechanism-proxy CV, and full heat-ablation robustness.\n')


@dataclass
class Encoders:
    supplier_ohe: Optional[OneHotEncoder]
    src_mlb: Optional[MultiLabelBinarizer]
    app_mlb: Optional[MultiLabelBinarizer]
    scalers: Dict[str, StandardScaler]


class DataBuilder:
    def __init__(self, excel_path: str, media_path: str, neo4j_dir: str, out_dir: Optional[str]=None):
        self.excel_path = excel_path
        self.media_path = media_path
        self.neo4j_dir = neo4j_dir
        self.out_dir = out_dir

    def load(self) -> pd.DataFrame:
        price = pd.read_excel(self.excel_path)
        price = price.copy()
        price['price'] = pd.to_numeric(price['price'], errors='coerce')
        price = price[price['price'].gt(0)].copy()
        price['supplier'] = price['supplier'].astype(str)
        price['name'] = price['name'].astype(str)
        price['y_ln'] = price['price'].apply(safe_ln)
        price = price.dropna(subset=['y_ln']).reset_index(drop=True)
        price['Total_Asset_num'] = price.get('Total_Asset', 0).apply(parse_asset_num)

        nodes_dp = pd.read_csv(os.path.join(self.neo4j_dir, 'nodes_dataproduct.csv'))
        rel_app = pd.read_csv(os.path.join(self.neo4j_dir, 'rel_applied_to.csv'))
        rel_src = pd.read_csv(os.path.join(self.neo4j_dir, 'rel_source_industry.csv'))
        nodes_dp = nodes_dp.rename(columns={'name_anon': 'name', 'supplier_anon': 'supplier', 'desc_anon': 'desc'})
        rel_app = rel_app.rename(columns={'app_name': 'app'})
        rel_src = rel_src.rename(columns={'src_name': 'src'})
        dp = nodes_dp[['dp_id', 'name', 'supplier', 'desc']].copy()
        dp['name'] = dp['name'].astype(str)
        dp['supplier'] = dp['supplier'].astype(str)

        app_map = rel_app.groupby('dp_id')['app'].apply(lambda s: sorted({str(v).strip() for v in s.dropna() if str(v).strip()})).to_dict()
        src_map = rel_src.groupby('dp_id')['src'].apply(lambda s: sorted({str(v).strip() for v in s.dropna() if str(v).strip()})).to_dict()
        dp['app_list'] = dp['dp_id'].map(app_map).apply(lambda x: x if isinstance(x, list) else [])
        dp['src_list'] = dp['dp_id'].map(src_map).apply(lambda x: x if isinstance(x, list) else [])
        dp = dp.sort_values('dp_id').drop_duplicates(['name', 'supplier'], keep='first')

        df = price.merge(dp[['name', 'supplier', 'desc', 'app_list', 'src_list']], on=['name', 'supplier'], how='left')
        df['desc'] = df['desc'].fillna('')
        df['app_list'] = df['app_list'].apply(lambda x: x if isinstance(x, list) else [])
        df['src_list'] = df['src_list'].apply(lambda x: x if isinstance(x, list) else [])

        # Product embeddings and competitor-neighborhood summaries.
        # Existing files are preserved; if they are missing, regenerate them from desc + KG app/src labels.
        data_dir = os.path.dirname(self.excel_path) or '.'
        ensure_product_reasoning_input_files(df, data_dir=data_dir, media_path=self.media_path, out_dir=self.out_dir, force=False)

        emb = pd.read_csv(os.path.join(data_dir, 'STEP0_product_textemb.csv'))
        emb['name'] = emb['name'].astype(str)
        emb['supplier'] = emb['supplier'].astype(str)
        df = df.merge(emb, on=['name', 'supplier'], how='left')

        comp = pd.read_csv(os.path.join(data_dir, 'STEP0_product_competitor_summary.csv'))
        comp['name'] = comp['name'].astype(str)
        comp['supplier'] = comp['supplier'].astype(str)
        df = df.merge(comp, on=['name', 'supplier'], how='left', suffixes=('', '_comp'))
        if 'prod_key' not in df.columns:
            df['prod_key'] = 'PROD::' + df['name'] + '||SUP::' + df['supplier']

        # Media heat
        # Consistent with Appendix D and the manuscript:
        # Heat_web/news/weixin are log-transformed public visibility counts,
        # and Heat_total is the sum of the three log-components.
        media = pd.read_excel(self.media_path)
        media = media.rename(columns={'keyword': 'app'})
        for c in ['sogou_web_results', 'sina_news_results', 'weixin_article_results']:
            media[c] = pd.to_numeric(media[c], errors='coerce').fillna(0.0)
        media['heat_web'] = np.log1p(media['sogou_web_results'])
        media['heat_news'] = np.log1p(media['sina_news_results'])
        media['heat_weixin'] = np.log1p(media['weixin_article_results'])
        media['heat_total'] = media['heat_web'] + media['heat_news'] + media['heat_weixin']
        app_heat = media.set_index('app')[['heat_total', 'heat_web', 'heat_news', 'heat_weixin']].to_dict('index')

        def app_heat_agg(apps: List[str], key: str, agg='mean') -> float:
            vals = [float(app_heat.get(a, {}).get(key, 0.0)) for a in apps]
            if not vals:
                return 0.0
            if agg == 'max':
                return float(np.max(vals))
            if agg == 'sum':
                return float(np.sum(vals))
            return float(np.mean(vals))

        for key in ['heat_total', 'heat_web', 'heat_news', 'heat_weixin']:
            df[f'demand_{key}_mean'] = df['app_list'].apply(lambda x: app_heat_agg(x, key, 'mean'))
        df['demand_heat_total_max'] = df['app_list'].apply(lambda x: app_heat_agg(x, 'heat_total', 'max'))
        df['demand_heat_total_sum'] = df['app_list'].apply(lambda x: app_heat_agg(x, 'heat_total', 'sum'))

        # Source-side supplier counts and application-side audit counts.
        # By definition in this manuscript, Nmarket is the source-side supplier participation
        # measure: the mean number of distinct suppliers across the product's associated
        # source-industry contexts. Application-side counts are retained only for audit or
        # descriptive checks and are not used to construct Nmarket, Bs, HHI_proxy, or T.
        app_edges = rel_app.merge(dp[['dp_id', 'supplier']], on='dp_id', how='left')
        src_edges = rel_src.merge(dp[['dp_id', 'supplier']], on='dp_id', how='left')
        app_supplier_cnt = app_edges.groupby('app')['supplier'].nunique().to_dict()
        src_supplier_cnt = src_edges.groupby('src')['supplier'].nunique().to_dict()
        src_product_cnt = src_edges.groupby('src')['dp_id'].nunique().to_dict()
        app_product_cnt = app_edges.groupby('app')['dp_id'].nunique().to_dict()

        def agg_from_map(labels, mp, agg='mean', default=0.0):
            vals = [float(mp.get(v, default)) for v in labels]
            if not vals:
                return 0.0
            if agg == 'max':
                return float(np.max(vals))
            if agg == 'sum':
                return float(np.sum(vals))
            return float(np.mean(vals))

        # Nmarket: source-side supplier participation.
        # For each product, compute the mean number of distinct suppliers across its
        # associated source-industry contexts. This is the supply-side base for the
        # paper-facing variables Bs, HHI_proxy, and T.
        df['supply_src_supplier_mean'] = df['src_list'].apply(lambda x: agg_from_map(x, src_supplier_cnt, 'mean'))

        # Application-side supplier/product counts are retained only for audit or
        # descriptive checks. They are deliberately not used in the main Supply, Demand,
        # or Market blocks and do not enter the construction of Nmarket, Bs, HHI_proxy, or T.
        df['market_app_supplier_mean'] = df['app_list'].apply(lambda x: agg_from_map(x, app_supplier_cnt, 'mean'))
        df['market_app_product_mean'] = df['app_list'].apply(lambda x: agg_from_map(x, app_product_cnt, 'mean'))

        # Source-side supply and market-structure controls.
        # Nmarket is the mean number of distinct suppliers in the product's source-industry contexts.
        # HHI_proxy = 1/Nmarket; Bs = ln(1 + Nmarket);
        # T = ln(1 + Nmarket * Dmedia_heat), where Dmedia_heat = exp(heat_total) - 1.
        nmarket = df['supply_src_supplier_mean'].astype(float).clip(lower=1.0)
        df['market_hhi_proxy_mean'] = 1.0 / nmarket
        df['bs'] = np.log1p(nmarket)
        dmedia_heat = np.expm1(df['demand_heat_total_mean'].astype(float)).clip(lower=0.0)
        df['market_thickness_mean'] = np.log1p(nmarket * dmedia_heat)

        # Paper-facing aliases used in Appendix D and in the variable-construction text.
        # Nmarket is intentionally retained as the source-side supplier participation variable.
        # These aliases do not create additional regressors; they make STEP0 output, Appendix D,
        # and the manuscript terminology consistent and directly auditable.
        df['heat_web'] = df['demand_heat_web_mean']
        df['heat_news'] = df['demand_heat_news_mean']
        df['heat_weixin'] = df['demand_heat_weixin_mean']
        df['heat_total'] = df['demand_heat_total_mean']
        df['Nmarket'] = nmarket
        df['Dmedia_heat'] = dmedia_heat
        df['HHI_proxy'] = df['market_hhi_proxy_mean']
        df['Bs'] = df['bs']
        df['T'] = df['market_thickness_mean']

        # Product block columns from precomputed embeddings + competitor summaries
        text_cols = [c for c in df.columns if c.startswith('textemb_')]
        comp_cols = [c for c in ['comp_app_entropy','comp_src_entropy','comp_app_top1_prob','comp_src_top1_prob','comp_expected_heat_total'] if c in df.columns]
        for c in text_cols + comp_cols:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)

        # Ensure lists are clean
        df['src_list'] = df['src_list'].apply(lambda x: x if isinstance(x, list) else parse_list_cell(x))
        df['app_list'] = df['app_list'].apply(lambda x: x if isinstance(x, list) else parse_list_cell(x))
        df['topic_id'] = '0'
        self.text_cols = text_cols
        self.comp_cols = comp_cols
        return df


def block_columns(df: pd.DataFrame, builder: Optional[DataBuilder]=None) -> Dict[str, List[str]]:
    text_cols = [c for c in df.columns if c.startswith("textemb_")]
    comp_cols = [c for c in ["comp_app_entropy","comp_src_entropy","comp_app_top1_prob","comp_src_top1_prob","comp_expected_heat_total"] if c in df.columns]
    product_cols = text_cols + comp_cols

    # Main Supply block: observable source-side supplier participation, defined in the
    # manuscript/Appendix D as Bs = ln(1 + Nmarket), where Nmarket is the mean number
    # of distinct suppliers across the product's associated source-industry contexts.
    # Bs is assigned to Supply rather than Market so that the empirical block structure
    # matches the model formula with a distinct supply term.
    supply_cols = [c for c in ['Bs'] if c in df.columns]

    # Main Demand block: one composite downstream demand-attention index, using the same
    # paper-facing variable name as Appendix D. Component heat variables are retained
    # in STEP0 and Appendix D for transparency only, not as separate model regressors.
    demand_cols = [c for c in ['heat_total'] if c in df.columns]

    # Main Market block: market concentration and demand-weighted market thickness.
    # HHI_proxy and T are constructed from the same source-side Nmarket base, with T
    # additionally incorporating downstream demand exposure. Bs is deliberately excluded
    # here because it is the Supply block variable.
    market_cols = [c for c in ['HHI_proxy','T'] if c in df.columns]

    return {'Product': product_cols, 'Supply': supply_cols, 'Demand': demand_cols, 'Market': market_cols}


def fit_block_encoders(train_df: pd.DataFrame, numeric_block_cols: Dict[str, List[str]]) -> Encoders:
    try:
        supplier_ohe = OneHotEncoder(handle_unknown='ignore', sparse_output=False)
    except TypeError:  # compatibility with older scikit-learn versions
        supplier_ohe = OneHotEncoder(handle_unknown='ignore', sparse=False)
    supplier_ohe.fit(train_df[['supplier']].astype(str))
    src_mlb = MultiLabelBinarizer()
    src_mlb.fit(train_df['src_list'])
    app_mlb = MultiLabelBinarizer()
    app_mlb.fit(train_df['app_list'])
    scalers = {}
    for block, cols in numeric_block_cols.items():
        if cols:
            sc = StandardScaler()
            tmp = train_df[cols].apply(pd.to_numeric, errors='coerce').fillna(0.0)
            sc.fit(tmp)
            scalers[block] = sc
    return Encoders(supplier_ohe, src_mlb, app_mlb, scalers)


def transform_blocks(df: pd.DataFrame, enc: Encoders, numeric_block_cols: Dict[str, List[str]]) -> Dict[str, np.ndarray]:
    out: Dict[str, np.ndarray] = {}
    out['Supplier FE'] = enc.supplier_ohe.transform(df[['supplier']].astype(str)) if enc.supplier_ohe else np.zeros((len(df), 0))
    out['Src FE'] = mlb_transform_no_warn(enc.src_mlb, df['src_list']) if enc.src_mlb else np.zeros((len(df), 0))
    out['App FE'] = mlb_transform_no_warn(enc.app_mlb, df['app_list']) if enc.app_mlb else np.zeros((len(df), 0))
    for block in ['Product','Supply','Demand','Market']:
        cols = numeric_block_cols.get(block, [])
        if cols:
            Xdf = df[cols].apply(pd.to_numeric, errors='coerce').fillna(0.0)
            out[block] = enc.scalers[block].transform(Xdf) if block in enc.scalers else Xdf.values
        else:
            out[block] = np.zeros((len(df), 0))
    return out


def join_blocks(block_mats: Dict[str, np.ndarray], chosen: List[str]) -> np.ndarray:
    if not chosen:
        return np.ones((next(iter(block_mats.values())).shape[0], 1))
    mats = [block_mats[b] for b in chosen]
    mats = [m for m in mats if m.ndim == 2 and m.shape[1] > 0]
    if not mats:
        return np.ones((next(iter(block_mats.values())).shape[0], 1))
    return np.concatenate(mats, axis=1)


def fit_predict(train_df: pd.DataFrame, test_df: pd.DataFrame, chosen: List[str], numeric_block_cols: Dict[str, List[str]]) -> Tuple[np.ndarray, np.ndarray]:
    enc = fit_block_encoders(train_df, numeric_block_cols)
    tr_blocks = transform_blocks(train_df, enc, numeric_block_cols)
    te_blocks = transform_blocks(test_df, enc, numeric_block_cols)
    Xtr = join_blocks(tr_blocks, chosen)
    Xte = join_blocks(te_blocks, chosen)
    ytr = train_df['y_ln'].values.astype(float)
    model = BayesianRidge()
    model.fit(Xtr, ytr)
    pred = model.predict(Xte)
    pred_tr = model.predict(Xtr)
    return pred, pred_tr


def fullsample_metrics(df: pd.DataFrame, chosen: List[str], numeric_block_cols: Dict[str, List[str]]) -> Dict[str, float]:
    pred, _ = fit_predict(df, df, chosen, numeric_block_cols)
    return metrics_on_ln_and_price(df['y_ln'].values, pred)


def cv_metrics(df: pd.DataFrame, splits: List[Tuple[np.ndarray,np.ndarray]], chosen: List[str], numeric_block_cols: Dict[str, List[str]]) -> Dict[str, float]:
    rows = []
    for tr, te in splits:
        train_df = df.iloc[tr].copy(); test_df = df.iloc[te].copy()
        pred, _ = fit_predict(train_df, test_df, chosen, numeric_block_cols)
        rows.append(metrics_on_ln_and_price(test_df['y_ln'].values, pred))
    out = {}
    for k in rows[0].keys():
        vals = [r[k] for r in rows]
        out[k] = float(np.mean(vals))
        out[k+'_SD'] = float(np.std(vals, ddof=0))
    return out


def scenario_subsets(blocks: List[str]) -> List[Tuple[str,...]]:
    out = [tuple()]
    for r in range(1, len(blocks)+1):
        out.extend(combinations(blocks, r))
    return out


def fullsample_r2_for_subset(df: pd.DataFrame, subset: Tuple[str,...], numeric_block_cols: Dict[str,List[str]]) -> float:
    return fullsample_metrics(df, list(subset), numeric_block_cols)['R2']


def shapley_fullsample(df: pd.DataFrame, blocks: List[str], numeric_block_cols: Dict[str,List[str]]) -> pd.DataFrame:
    subsets = scenario_subsets(blocks)
    cache = {s: fullsample_r2_for_subset(df, s, numeric_block_cols) for s in subsets}
    rows = []
    n = len(blocks)
    fac = math.factorial
    for b in blocks:
        phi = 0.0
        others = [x for x in blocks if x != b]
        for r in range(len(others)+1):
            for S in combinations(others, r):
                S = tuple(S)
                Sw = tuple(list(S)+[b])
                w = fac(len(S))*fac(n-len(S)-1)/fac(n)
                phi += w * (cache[tuple(sorted(Sw, key=blocks.index))] - cache[S])
        rows.append({'component': b, 'Shapley': float(phi)})
    return pd.DataFrame(rows)



def _safe_pearson(x: np.ndarray, y: np.ndarray) -> float:
    x, y = np.asarray(x, dtype=float), np.asarray(y, dtype=float)
    m = np.isfinite(x) & np.isfinite(y)
    x, y = x[m], y[m]
    if len(x) < 3 or np.std(x) < 1e-12 or np.std(y) < 1e-12:
        return np.nan
    return float(np.corrcoef(x, y)[0, 1])


def _safe_spearman(x: np.ndarray, y: np.ndarray) -> float:
    x, y = np.asarray(x, dtype=float), np.asarray(y, dtype=float)
    m = np.isfinite(x) & np.isfinite(y)
    x, y = x[m], y[m]
    if len(x) < 3 or np.std(x) < 1e-12 or np.std(y) < 1e-12:
        return np.nan
    return spearman_rho(x, y)


def _safe_slope(x: np.ndarray, y: np.ndarray) -> float:
    x, y = np.asarray(x, dtype=float), np.asarray(y, dtype=float)
    m = np.isfinite(x) & np.isfinite(y)
    x, y = x[m], y[m]
    if len(x) < 3 or np.std(x) < 1e-12 or np.std(y) < 1e-12:
        return np.nan
    return float(LinearRegression().fit(x.reshape(-1, 1), y).coef_[0])


def _heat_ablation_row(level: str, method: str, d: pd.DataFrame, structural_col: str, heat_col: str, y_col: str = 'y_ln') -> Dict[str, Any]:
    d = d[[structural_col, heat_col, y_col]].apply(pd.to_numeric, errors='coerce').dropna()
    if len(d):
        structural_mean = float(d[structural_col].mean())
        heat_mean = float(d[heat_col].mean())
        delta_mean = float((d[heat_col] - d[structural_col]).mean())
    else:
        structural_mean = np.nan
        heat_mean = np.nan
        delta_mean = np.nan
    return {
        'Entity level': level,
        'Method': method,
        'N': int(len(d)),
        'Structural variable': structural_col,
        'Heat-biased variable': heat_col,
        'Structural mean': structural_mean,
        'Heat-biased mean': heat_mean,
        'Mean Δ': delta_mean,
        'Consistency Pearson r': _safe_pearson(d[structural_col].values, d[heat_col].values) if len(d) else np.nan,
        'Consistency Spearman ρ': _safe_spearman(d[structural_col].values, d[heat_col].values) if len(d) else np.nan,
        'Structural assoc. Spearman ρ': _safe_spearman(d[structural_col].values, d[y_col].values) if len(d) else np.nan,
        'Structural assoc. slope': _safe_slope(d[structural_col].values, d[y_col].values) if len(d) else np.nan,
        'Heat-biased assoc. Spearman ρ': _safe_spearman(d[heat_col].values, d[y_col].values) if len(d) else np.nan,
        'Heat-biased assoc. slope': _safe_slope(d[heat_col].values, d[y_col].values) if len(d) else np.nan,
    }


def _compact_heat_ablation_table(e3_raw: pd.DataFrame) -> pd.DataFrame:
    """Compact Appendix E.3 for manuscript use.

    The raw heat-ablation helper keeps variable names for auditability, but the appendix table
    should focus on the substantive comparison: signal level, structural/heat-biased
    consistency, and association with ln(price).
    """
    cols = [
        'Entity level', 'Method', 'N',
        'Structural mean', 'Heat-biased mean', 'Mean Δ',
        'Consistency (r / ρ)',
        'Structural assoc. (ρ / slope)',
        'Heat-biased assoc. (ρ / slope)',
    ]
    if e3_raw is None or e3_raw.empty:
        return pd.DataFrame(columns=cols)
    out = e3_raw.copy()
    out['Consistency (r / ρ)'] = out.apply(
        lambda r: f"{_fmt_num(r.get('Consistency Pearson r'))} / {_fmt_num(r.get('Consistency Spearman ρ'))}", axis=1
    )
    out['Structural assoc. (ρ / slope)'] = out.apply(
        lambda r: f"{_fmt_num(r.get('Structural assoc. Spearman ρ'))} / {_fmt_num(r.get('Structural assoc. slope'))}", axis=1
    )
    out['Heat-biased assoc. (ρ / slope)'] = out.apply(
        lambda r: f"{_fmt_num(r.get('Heat-biased assoc. Spearman ρ'))} / {_fmt_num(r.get('Heat-biased assoc. slope'))}", axis=1
    )
    for c in ['Structural mean', 'Heat-biased mean', 'Mean Δ']:
        if c in out.columns:
            out[c] = out[c].apply(lambda x: _fmt_num(x))
    return out[cols]


def build_table7_ppr_mppr(df: pd.DataFrame, out_dir: str) -> pd.DataFrame:
    """Main-text Table 7: product-level PPR/MPPR structural vs heat-biased propagation."""
    base = df[['prod_key', 'y_ln']].drop_duplicates('prod_key').copy()
    rows = []
    ppr_path = os.path.join(out_dir, 'STEP2_KG_reasoning_product_ppr.csv')
    mppr_path = os.path.join(out_dir, 'STEP2_KG_reasoning_product_mppr.csv')
    if os.path.exists(ppr_path):
        ppr = base.merge(pd.read_csv(ppr_path), on='prod_key', how='inner')
        if {'prod_ppr_expected_heat_total_struct', 'prod_ppr_expected_heat_total_tele'}.issubset(ppr.columns):
            rows.append(_heat_ablation_row('Product', 'PPR', ppr, 'prod_ppr_expected_heat_total_struct', 'prod_ppr_expected_heat_total_tele'))
    if os.path.exists(mppr_path):
        mppr = base.merge(pd.read_csv(mppr_path), on='prod_key', how='inner')
        if {'prod_mp_ppr_expected_heat_total_struct', 'prod_mp_ppr_expected_heat_total_tele'}.issubset(mppr.columns):
            rows.append(_heat_ablation_row('Product', 'MPPR', mppr, 'prod_mp_ppr_expected_heat_total_struct', 'prod_mp_ppr_expected_heat_total_tele'))
    out = pd.DataFrame(rows)
    if out.empty:
        return pd.DataFrame(columns=['Method','N','Consistency (r / ρ)','Structural assoc. (ρ / slope)','Heat-biased assoc. (ρ / slope)'])
    out['Consistency (r / ρ)'] = out.apply(lambda r: f"{_fmt_num(r['Consistency Pearson r'])} / {_fmt_num(r['Consistency Spearman ρ'])}", axis=1)
    out['Structural assoc. (ρ / slope)'] = out.apply(lambda r: f"{_fmt_num(r['Structural assoc. Spearman ρ'])} / {_fmt_num(r['Structural assoc. slope'])}", axis=1)
    out['Heat-biased assoc. (ρ / slope)'] = out.apply(lambda r: f"{_fmt_num(r['Heat-biased assoc. Spearman ρ'])} / {_fmt_num(r['Heat-biased assoc. slope'])}", axis=1)
    return out[['Method','N','Consistency (r / ρ)','Structural assoc. (ρ / slope)','Heat-biased assoc. (ρ / slope)']]


def _mechanism_cv(level: str, method: str, d: pd.DataFrame, cols: List[str]) -> Dict[str, Any]:
    d = d.copy()
    cols = [c for c in cols if c in d.columns]
    if 'y_ln' not in d.columns or not cols or len(d) < 5:
        return {'Entity level': level, 'Method': method, 'N': int(len(d)), 'Features': '; '.join(cols), 'R2': np.nan, 'MAE': np.nan, 'RMSE': np.nan}
    d[cols] = d[cols].apply(pd.to_numeric, errors='coerce').fillna(0.0)
    d['y_ln'] = pd.to_numeric(d['y_ln'], errors='coerce')
    d = d.dropna(subset=['y_ln']).reset_index(drop=True)
    if len(d) < 5:
        return {'Entity level': level, 'Method': method, 'N': int(len(d)), 'Features': '; '.join(cols), 'R2': np.nan, 'MAE': np.nan, 'RMSE': np.nan}
    mets = []
    for tr, te in KFold(n_splits=min(5, len(d)), shuffle=True, random_state=RANDOM_STATE).split(d):
        sc = StandardScaler().fit(d.iloc[tr][cols])
        Xtr = sc.transform(d.iloc[tr][cols]); Xte = sc.transform(d.iloc[te][cols])
        ytr = d.iloc[tr]['y_ln'].values; yte = d.iloc[te]['y_ln'].values
        pred = BayesianRidge().fit(Xtr, ytr).predict(Xte)
        mets.append(metrics_on_ln_and_price(yte, pred))
    return {'Entity level': level, 'Method': method, 'N': int(len(d)), 'Features': '; '.join(cols), 'R2': float(np.mean([m['R2'] for m in mets])), 'MAE': float(np.mean([m['MAE'] for m in mets])), 'RMSE': float(np.mean([m['RMSE'] for m in mets]))}


def build_appendix_e_tables(df: pd.DataFrame, out_dir: str) -> Dict[str, pd.DataFrame]:
    """Appendix E: PPR/MPPR definitions, mechanism-proxy CV, and full heat-ablation robustness."""
    e1 = pd.DataFrame([
        {'Variable': 'prod_ppr_expected_heat_total_struct', 'Method': 'PPR', 'Entity level': 'Product', 'Definition': 'Topology-only product-to-application demand exposure from heterogeneous-graph PPR.'},
        {'Variable': 'prod_ppr_expected_heat_total_tele', 'Method': 'PPR', 'Entity level': 'Product', 'Definition': 'Heat-teleport product-to-application demand exposure from heterogeneous-graph PPR.'},
        {'Variable': 'prod_ppr_app_entropy', 'Method': 'PPR', 'Entity level': 'Product', 'Definition': 'Entropy of the product-level PPR application posterior.'},
        {'Variable': 'prod_ppr_src_entropy', 'Method': 'PPR', 'Entity level': 'Product', 'Definition': 'Entropy of the product-level PPR source-industry posterior.'},
        {'Variable': 'prod_ppr_app_top1_prob / prod_ppr_src_top1_prob', 'Method': 'PPR', 'Entity level': 'Product', 'Definition': 'Top-1 concentration probabilities for application and source PPR posteriors.'},
        {'Variable': 'prod_mp_ppr_expected_heat_total_struct', 'Method': 'MPPR', 'Entity level': 'Product', 'Definition': 'Topology-only product-level demand exposure aggregated from meta-path-constrained PPR posteriors.'},
        {'Variable': 'prod_mp_ppr_expected_heat_total_tele', 'Method': 'MPPR', 'Entity level': 'Product', 'Definition': 'Heat-teleport product-level demand exposure aggregated from meta-path-constrained PPR posteriors.'},
        {'Variable': 'prod_mp_ppr_app_entropy', 'Method': 'MPPR', 'Entity level': 'Product', 'Definition': 'Entropy of the product-level MPPR application posterior.'},
        {'Variable': 'prod_mp_ppr_src_entropy', 'Method': 'MPPR', 'Entity level': 'Product', 'Definition': 'Entropy of the product-level MPPR source-industry posterior.'},
        {'Variable': 'prod_mp_ppr_app_top1_prob / prod_mp_ppr_src_top1_prob', 'Method': 'MPPR', 'Entity level': 'Product', 'Definition': 'Top-1 concentration probabilities for application and source MPPR posteriors.'},
    ])

    e2_rows: List[Dict[str, Any]] = []
    base_prod = df[['prod_key','y_ln']].drop_duplicates('prod_key').copy()
    ppr_path = os.path.join(out_dir, 'STEP2_KG_reasoning_product_ppr.csv')
    mppr_path = os.path.join(out_dir, 'STEP2_KG_reasoning_product_mppr.csv')
    if os.path.exists(ppr_path):
        ppr = base_prod.merge(pd.read_csv(ppr_path), on='prod_key', how='inner')
        e2_rows.append(_mechanism_cv('Product', 'PPR', ppr, ['prod_ppr_expected_heat_total','prod_ppr_app_entropy','prod_ppr_src_entropy','prod_ppr_app_top1_prob','prod_ppr_src_top1_prob']))
    if os.path.exists(mppr_path):
        mppr = base_prod.merge(pd.read_csv(mppr_path), on='prod_key', how='inner')
        e2_rows.append(_mechanism_cv('Product', 'MPPR', mppr, ['prod_mp_ppr_expected_heat_total','prod_mp_ppr_app_entropy','prod_mp_ppr_src_entropy','prod_mp_ppr_app_top1_prob','prod_mp_ppr_src_top1_prob']))

    sup_path = os.path.join(out_dir, 'STEP2_mechanisms_supplier.csv')
    if os.path.exists(sup_path):
        sup_y = df.groupby('supplier', as_index=False).agg(y_ln=('y_ln','mean'))
        sup = sup_y.merge(pd.read_csv(sup_path), on='supplier', how='inner')
        e2_rows.append(_mechanism_cv('Supplier', 'PPR', sup, ['ppr_expected_heat_total','ppr_app_entropy','ppr_src_entropy','ppr_app_top1_prob','ppr_src_top1_prob']))
        e2_rows.append(_mechanism_cv('Supplier', 'MPPR', sup, ['mp_ppr_expected_heat_total','mp_ppr_app_entropy','mp_ppr_src_entropy','mp_ppr_app_top1_prob','mp_ppr_src_top1_prob']))

    src_path = os.path.join(out_dir, 'STEP2_mechanisms_src.csv')
    if os.path.exists(src_path):
        src_y = df.explode('src_list').groupby('src_list', as_index=False).agg(y_ln=('y_ln','mean')).rename(columns={'src_list':'src'})
        srcm = src_y.merge(pd.read_csv(src_path), on='src', how='inner')
        e2_rows.append(_mechanism_cv('Src', 'PPR', srcm, ['ppr_expected_heat_total','ppr_app_entropy','ppr_app_top1_prob']))
        e2_rows.append(_mechanism_cv('Src', 'MPPR', srcm, ['mp_ppr_expected_heat_total','mp_ppr_app_entropy','mp_ppr_app_top1_prob']))

    app_path = os.path.join(out_dir, 'STEP2_mechanisms_app.csv')
    if os.path.exists(app_path):
        app_y = df.explode('app_list').groupby('app_list', as_index=False).agg(y_ln=('y_ln','mean')).rename(columns={'app_list':'app'})
        appm = app_y.merge(pd.read_csv(app_path), on='app', how='inner')
        e2_rows.append(_mechanism_cv('App', 'PPR', appm, ['ppr_src_entropy','ppr_src_top1_prob','heat_total']))
        e2_rows.append(_mechanism_cv('App', 'MPPR', appm, ['mp_ppr_src_entropy','mp_ppr_sup_entropy','mp_ppr_src_top1_prob','mp_ppr_sup_top1_prob','heat_total']))

    e2 = pd.DataFrame(e2_rows)

    e3_rows: List[Dict[str, Any]] = []
    if os.path.exists(ppr_path):
        ppr = base_prod.merge(pd.read_csv(ppr_path), on='prod_key', how='inner')
        if {'prod_ppr_expected_heat_total_struct','prod_ppr_expected_heat_total_tele'}.issubset(ppr.columns):
            e3_rows.append(_heat_ablation_row('Product', 'PPR', ppr, 'prod_ppr_expected_heat_total_struct', 'prod_ppr_expected_heat_total_tele'))
    if os.path.exists(mppr_path):
        mppr = base_prod.merge(pd.read_csv(mppr_path), on='prod_key', how='inner')
        if {'prod_mp_ppr_expected_heat_total_struct','prod_mp_ppr_expected_heat_total_tele'}.issubset(mppr.columns):
            e3_rows.append(_heat_ablation_row('Product', 'MPPR', mppr, 'prod_mp_ppr_expected_heat_total_struct', 'prod_mp_ppr_expected_heat_total_tele'))
    if os.path.exists(sup_path):
        sup_y = df.groupby('supplier', as_index=False).agg(y_ln=('y_ln','mean'))
        sup = sup_y.merge(pd.read_csv(sup_path), on='supplier', how='inner')
        if {'ppr_expected_heat_total_struct','ppr_expected_heat_total_tele'}.issubset(sup.columns):
            e3_rows.append(_heat_ablation_row('Supplier', 'PPR', sup, 'ppr_expected_heat_total_struct', 'ppr_expected_heat_total_tele'))
        if {'mp_ppr_expected_heat_total_struct','mp_ppr_expected_heat_total_tele'}.issubset(sup.columns):
            e3_rows.append(_heat_ablation_row('Supplier', 'MPPR', sup, 'mp_ppr_expected_heat_total_struct', 'mp_ppr_expected_heat_total_tele'))
    if os.path.exists(src_path):
        src_y = df.explode('src_list').groupby('src_list', as_index=False).agg(y_ln=('y_ln','mean')).rename(columns={'src_list':'src'})
        srcm = src_y.merge(pd.read_csv(src_path), on='src', how='inner')
        if {'ppr_expected_heat_total_struct','ppr_expected_heat_total_tele'}.issubset(srcm.columns):
            e3_rows.append(_heat_ablation_row('Src', 'PPR', srcm, 'ppr_expected_heat_total_struct', 'ppr_expected_heat_total_tele'))
        if {'mp_ppr_expected_heat_total_struct','mp_ppr_expected_heat_total_tele'}.issubset(srcm.columns):
            e3_rows.append(_heat_ablation_row('Src', 'MPPR', srcm, 'mp_ppr_expected_heat_total_struct', 'mp_ppr_expected_heat_total_tele'))
    e3_raw = pd.DataFrame(e3_rows)
    # Keep the appendix-facing table compact; the raw rows with variable names can be
    # regenerated from STEP2_* files and are not needed in the manuscript table.
    e3 = _compact_heat_ablation_table(e3_raw)

    return {'Appendix_E1': e1, 'Appendix_E2': e2, 'Appendix_E3': e3}

def build_tables(df: pd.DataFrame, out_dir: str) -> Dict[str,pd.DataFrame]:
    ensure_dir(out_dir)
    nb = block_columns(df)
    print('[build_tables] Table 1', flush=True)
    # Table1
    t1 = pd.DataFrame([
        {'Variable':'price','Obs':len(df),'Mean':df['price'].mean(),'SD':df['price'].std(ddof=1),'Min':df['price'].min(),'P25':df['price'].quantile(0.25),'P50':df['price'].quantile(0.5),'P75':df['price'].quantile(0.75),'Max':df['price'].max()},
        {'Variable':'ln(price)','Obs':len(df),'Mean':df['y_ln'].mean(),'SD':df['y_ln'].std(ddof=1),'Min':df['y_ln'].min(),'P25':df['y_ln'].quantile(0.25),'P50':df['y_ln'].quantile(0.5),'P75':df['y_ln'].quantile(0.75),'Max':df['y_ln'].max()},
    ])

    # Splits
    kf = KFold(n_splits=N_SPLITS, shuffle=True, random_state=RANDOM_STATE)
    random_splits = [(tr, te) for tr, te in kf.split(df)]
    groups = df['supplier'].astype(str).values
    gkf = GroupKFold(n_splits=min(N_SPLITS, len(np.unique(groups))))
    group_splits = [(tr, te) for tr, te in gkf.split(df, df['y_ln'].values, groups)]

    print('[build_tables] Table 2', flush=True)
    # Table2 Panel A independent full-sample fit
    rows = []
    rows.append({'Panel':'Panel A. Independent full-sample fit','Component':'Intercept only','Scenario':'None', **fullsample_metrics(df, [], nb)})
    for b in BLOCKS:
        rows.append({'Panel':'Panel A. Independent full-sample fit','Component':b,'Scenario':BLOCK_SHORT[b], **fullsample_metrics(df, [b], nb)})

    # Table2 Panel B conditional specifications.
    # "Base" denotes the supplier fixed-effect benchmark.
    # The first four rows are nested baseline expansions.
    # Supply, Demand, and Market are then added separately to the same
    # Base + Src + App + Product benchmark, followed by the full model.
    # This avoids the misleading impression that the Demand row necessarily
    # contains Supply, or that the Market row necessarily contains both Supply and Demand.
    base_core = ['Supplier FE']
    structural_base = ['Supplier FE', 'Src FE', 'App FE', 'Product']
    conditional_specs = [
        ('Supplier FE', base_core, 'Supplier FE'),
        ('Src FE', ['Supplier FE', 'Src FE'], 'Supplier FE + Src FE'),
        ('App FE', ['Supplier FE', 'Src FE', 'App FE'], 'Supplier FE + Src FE+ App FE'),
        ('Product', structural_base, 'Supplier FE + Src FE + App FE + Product'),
        ('Supply', structural_base + ['Supply'], 'Supplier FE + Src FE + App FE + Product + Supply'),
        ('Demand', structural_base + ['Demand'], 'Supplier FE + Src FE + App FE + Product + Demand'),
        ('Market', structural_base + ['Market'], 'Supplier FE + Src FE + App FE + Product + Market'),
        ('Full model', structural_base + ['Supply', 'Demand', 'Market'], 'Full: Supplier FE + Src FE + App FE + Product + Supply + Demand + Market'),
    ]
    for comp, chosen, scenario in conditional_specs:
        rows.append({'Panel':'Panel B. Conditional specifications','Component':comp,'Scenario':scenario, **fullsample_metrics(df, chosen, nb)})
    t2 = pd.DataFrame(rows)

    print('[build_tables] Table 3', flush=True)
    # Table3 full-sample ablation/drop-one/shapley
    rows = []
    for b in BLOCKS:
        rows.append({'Panel':'Panel A. Single-block full-sample fit','Component':b, **fullsample_metrics(df, [b], nb)})
    full_r2 = fullsample_metrics(df, BLOCKS, nb)['R2']
    for b in BLOCKS:
        chosen = [x for x in BLOCKS if x != b]
        met = fullsample_metrics(df, chosen, nb)
        rows.append({'Panel':'Panel B. Drop-one from full model','Component':b, 'R2':met['R2'], 'Delta_R2': full_r2 - met['R2']})
    shap = shapley_fullsample(df, BLOCKS, nb)
    for _, r in shap.iterrows():
        rows.append({'Panel':'Panel C. Shapley (full sample)','Component':r['component'], 'Shapley R2':r['Shapley']})
    t3 = pd.DataFrame(rows)

    print('[build_tables] Table 4', flush=True)
    # Table4 KFold
    rows = []
    rows.append({'Panel':'Panel A. Independent K-fold fit','Component':'Intercept only','Scenario':'None', **cv_metrics(df, random_splits, [], nb)})
    for b in BLOCKS:
        rows.append({'Panel':'Panel A. Independent K-fold fit','Component':b,'Scenario':BLOCK_SHORT[b], **cv_metrics(df, random_splits, [b], nb)})
    cum = []
    for b in BLOCKS:
        cum = cum + [b]
        rows.append({'Panel':'Panel B. Conditional additions (K-fold)','Component':b,'Scenario':' + '.join(cum), **cv_metrics(df, random_splits, cum, nb)})
    t4 = pd.DataFrame(rows)

    print('[build_tables] Table 5', flush=True)
    # Table5 Group KFold boundary
    rows = []
    cum = []
    for b in BLOCKS:
        cum = cum + [b]
        met_r = cv_metrics(df, random_splits, cum, nb)
        met_g = cv_metrics(df, group_splits, cum, nb)
        rows.append({'Component':b,'Scenario':' + '.join(cum),'Random K-fold R2':met_r['R2'],'Random K-fold SD':met_r['R2_SD'],'Group K-fold R2':met_g['R2'],'Group K-fold SD':met_g['R2_SD'],'Delta (Group-Random)':met_g['R2']-met_r['R2'],'Group MAE':met_g['MAE'],'Group RMSE':met_g['RMSE']})
    t5 = pd.DataFrame(rows)

    print('[build_tables] Table 6', flush=True)
    # Table6 mechanism proxies across entity levels
    rows = []
    # Supplier entity. Count variables are log-transformed for consistency with the market-structure controls.
    sup = df.groupby('supplier').agg(y_ln=('y_ln','mean'), n_prod_raw=('name','count'), demand=('heat_total','mean'), hhi=('HHI_proxy','mean'), thick=('T','mean')).reset_index()
    sup['n_prod'] = np.log1p(sup['n_prod_raw'])
    src_ent = df.explode('src_list').groupby('src_list').agg(y_ln=('y_ln','mean'), supplier_n_raw=('supplier','nunique'), demand=('heat_total','mean'), prod_n_raw=('name','count')).reset_index()
    src_ent['supplier_n'] = np.log1p(src_ent['supplier_n_raw'])
    src_ent['prod_n'] = np.log1p(src_ent['prod_n_raw'])
    app_ent = df.explode('app_list').groupby('app_list').agg(y_ln=('y_ln','mean'), supplier_n_raw=('supplier','nunique'), heat=('heat_total','mean'), prod_n_raw=('name','count')).reset_index()
    app_ent['supplier_n'] = np.log1p(app_ent['supplier_n_raw'])
    app_ent['prod_n'] = np.log1p(app_ent['prod_n_raw'])
    for ent, d, cols in [
        ('Supplier', sup, ['n_prod','demand','hhi','thick']),
        ('Src', src_ent, ['supplier_n','demand','prod_n']),
        ('App', app_ent, ['supplier_n','heat','prod_n']),
        ('Product', df[['prod_key','y_ln','comp_expected_heat_total','comp_app_entropy','comp_src_entropy','comp_app_top1_prob','comp_src_top1_prob']].drop_duplicates('prod_key'), ['comp_expected_heat_total','comp_app_entropy','comp_src_entropy','comp_app_top1_prob','comp_src_top1_prob']),
    ]:
        d = d.copy().fillna(0)
        if len(d) < 5:
            rows.append({'Entity level': ent, 'R2': np.nan, 'MAE': np.nan, 'RMSE': np.nan})
            continue
        splits = [(tr, te) for tr, te in KFold(n_splits=min(5, len(d)), shuffle=True, random_state=RANDOM_STATE).split(d)]
        mets = []
        for tr, te in splits:
            sc = StandardScaler().fit(d.iloc[tr][cols])
            Xtr = sc.transform(d.iloc[tr][cols]); Xte = sc.transform(d.iloc[te][cols])
            ytr = d.iloc[tr]['y_ln'].values; yte = d.iloc[te]['y_ln'].values
            m = BayesianRidge().fit(Xtr, ytr)
            pred = m.predict(Xte)
            mets.append(metrics_on_ln_and_price(yte, pred))
        rows.append({'Entity level': ent, 'R2': float(np.mean([m['R2'] for m in mets])), 'MAE': float(np.mean([m['MAE'] for m in mets])), 'RMSE': float(np.mean([m['RMSE'] for m in mets]))})
    t6 = pd.DataFrame(rows)

    print('[build_tables] Table 7', flush=True)
    # Table 7: true PPR/MPPR structural-vs-heat-biased propagation robustness.
    # It consumes the STEP2_KG_reasoning_product_ppr/mppr files generated in STEP0b.
    t7 = build_table7_ppr_mppr(df, out_dir)

    print('[build_tables] Appendix E', flush=True)
    appendix_e = build_appendix_e_tables(df, out_dir)

    tables = {'Table1':t1,'Table2':t2,'Table3':t3,'Table4':t4,'Table5':t5,'Table6':t6,'Table7':t7}
    tables.update(appendix_e)
    return tables


def _fmt_num(x: Any, decimals: int = 3) -> str:
    """Format a scalar for manuscript-style table output."""
    try:
        if pd.isna(x):
            return ''
        return f"{float(x):.{decimals}f}"
    except Exception:
        return '' if x is None else str(x)


def _fmt_mean_sd(row: pd.Series, mean_col: str, sd_col: str, decimals: int = 3) -> str:
    if mean_col not in row or pd.isna(row.get(mean_col)):
        return ''
    mean = _fmt_num(row.get(mean_col), decimals)
    if sd_col in row and not pd.isna(row.get(sd_col)):
        return f"{mean} ({_fmt_num(row.get(sd_col), decimals)})"
    return mean


def _panel_row(title: str, columns: List[str]) -> Dict[str, str]:
    row = {c: '' for c in columns}
    row[columns[0]] = title
    return row


def manuscript_table(table_name: str, df: pd.DataFrame, decimals: int = 3) -> pd.DataFrame:
    """Convert raw computational tables into manuscript-facing table layouts.

    The raw tables retain all numeric fields during computation. This function only
    changes the exported table layout: panel labels become standalone rows, K-fold
    standard deviations are placed in parentheses, and column names match the paper.
    """
    df = df.copy()
    if table_name == 'Table1':
        cols = ['Variable','Obs','Mean','SD','Min','P25','P50','P75','Max']
        out = df[cols].copy()
        for c in ['Mean','SD','Min','P25','P50','P75','Max']:
            out[c] = out[c].apply(lambda x: _fmt_num(x, decimals))
        return out

    if table_name == 'Table2':
        cols = ['Component','Model specification','R²','MAE','RMSE','MAPE','SMAPE']
        rows = []
        for panel, sub in df.groupby('Panel', sort=False):
            rows.append(_panel_row(panel, cols))
            for _, r in sub.iterrows():
                rows.append({
                    'Component': r.get('Component',''),
                    'Model specification': r.get('Scenario',''),
                    'R²': _fmt_num(r.get('R2'), decimals),
                    'MAE': _fmt_num(r.get('MAE'), decimals),
                    'RMSE': _fmt_num(r.get('RMSE'), decimals),
                    'MAPE': _fmt_num(r.get('MAPE'), decimals),
                    'SMAPE': _fmt_num(r.get('SMAPE'), decimals),
                })
        return pd.DataFrame(rows, columns=cols)

    if table_name == 'Table3':
        # Table 3 is a block-decomposition table, not a general error-metric table.
        # Therefore, it reports only R²-based explanatory contributions:
        # Panel A: single-block full-sample R²;
        # Panel B: R² after dropping each component from the full model and the corresponding ΔR²;
        # Panel C: Shapley R² contribution.
        # MAE, RMSE, MAPE, and SMAPE are intentionally omitted from Table 3 and remain reported
        # in the performance-oriented tables (Tables 2, 4, and 5).
        cols = ['Component','Single-block R²','R² without component','ΔR²','Shapley R² contribution']
        rows = []
        for panel, sub in df.groupby('Panel', sort=False):
            rows.append(_panel_row(panel, cols))
            for _, r in sub.iterrows():
                row = {c: '' for c in cols}
                row['Component'] = r.get('Component','')
                if str(panel).startswith('Panel A.'):
                    row['Single-block R²'] = _fmt_num(r.get('R2'), decimals)
                elif str(panel).startswith('Panel B.'):
                    row['R² without component'] = _fmt_num(r.get('R2'), decimals)
                    row['ΔR²'] = _fmt_num(r.get('Delta_R2'), decimals)
                elif str(panel).startswith('Panel C.'):
                    row['Shapley R² contribution'] = _fmt_num(r.get('Shapley R2'), decimals)
                else:
                    row['Single-block R²'] = _fmt_num(r.get('R2'), decimals)
                    row['ΔR²'] = _fmt_num(r.get('Delta_R2'), decimals)
                rows.append(row)
        return pd.DataFrame(rows, columns=cols)

    if table_name == 'Table4':
        cols = ['Component','Scenario','R²','MAE','RMSE','MAPE','SMAPE']
        rows = []
        for panel, sub in df.groupby('Panel', sort=False):
            rows.append(_panel_row(panel, cols))
            for _, r in sub.iterrows():
                rows.append({
                    'Component': r.get('Component',''),
                    'Scenario': r.get('Scenario',''),
                    'R²': _fmt_mean_sd(r, 'R2', 'R2_SD', decimals),
                    'MAE': _fmt_mean_sd(r, 'MAE', 'MAE_SD', decimals),
                    'RMSE': _fmt_mean_sd(r, 'RMSE', 'RMSE_SD', decimals),
                    'MAPE': _fmt_mean_sd(r, 'MAPE', 'MAPE_SD', decimals),
                    'SMAPE': _fmt_mean_sd(r, 'SMAPE', 'SMAPE_SD', decimals),
                })
        return pd.DataFrame(rows, columns=cols)

    if table_name == 'Table5':
        cols = ['Component','Scenario','Random K-fold R²','Group K-fold R²','Delta (Group–Random)','Group MAE','Group RMSE']
        rows = []
        for _, r in df.iterrows():
            rows.append({
                'Component': r.get('Component',''),
                'Scenario': r.get('Scenario',''),
                'Random K-fold R²': _fmt_mean_sd(r, 'Random K-fold R2', 'Random K-fold SD', decimals),
                'Group K-fold R²': _fmt_mean_sd(r, 'Group K-fold R2', 'Group K-fold SD', decimals),
                'Delta (Group–Random)': _fmt_num(r.get('Delta (Group-Random)'), decimals),
                'Group MAE': _fmt_num(r.get('Group MAE'), decimals),
                'Group RMSE': _fmt_num(r.get('Group RMSE'), decimals),
            })
        return pd.DataFrame(rows, columns=cols)

    if table_name == 'Table6':
        cols = ['Entity level','R²','MAE','RMSE']
        out = df.copy().rename(columns={'R2':'R²'})
        for c in ['R²','MAE','RMSE']:
            if c in out.columns:
                out[c] = out[c].apply(lambda x: _fmt_num(x, decimals))
        return out[cols]

    if table_name == 'Table7':
        return df.copy()

    if table_name in {'Appendix_E2'}:
        out = df.copy().rename(columns={'R2':'R²'})
        for c in ['R²','MAE','RMSE']:
            if c in out.columns:
                out[c] = out[c].apply(lambda x: _fmt_num(x, decimals))
        return out

    if table_name in {'Appendix_E3'}:
        # Appendix E.3 is already compacted in build_appendix_e_tables().
        return df.copy()

    return df


def style_ws(ws):
    header_fill = PatternFill(fill_type='solid', fgColor='D9E2F3')
    panel_fill = PatternFill(fill_type='solid', fgColor='EAF2F8')
    thin = Side(style='thin', color='BFBFBF')
    bottom = Border(bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill
        cell.border = bottom
    ws.freeze_panes = 'A2'
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=2):
        first = row[0].value
        if isinstance(first, str) and first.startswith('Panel '):
            for cell in row[:max_col]:
                cell.font = Font(bold=True, italic=True)
                cell.fill = panel_fill
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = bottom
        else:
            for cell in row[:max_col]:
                cell.alignment = Alignment(horizontal='center' if cell.column > 2 else 'left', vertical='center', wrap_text=True)
    for col in ws.columns:
        mx = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(mx+2, 10), 55)


def export_tables_1_to_11_and_appendices_1_to_5(out_dir: str, decimals: int = 3, excel_name: str = 'Paper_All_Tables_Merged.xlsx', csv_subdir: str = 'paper_tables_csv') -> str:
    csv_dir = os.path.join(out_dir, 'table_out', csv_subdir)
    ensure_dir(csv_dir)
    excel_path = os.path.join(out_dir, 'table_merge', excel_name)
    ensure_dir(os.path.dirname(excel_path))
    print('[export] write workbook', flush=True)
    tables = read_json(os.path.join(out_dir, 'table_manifest.json'))
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        for sheet, csvname in tables.items():
            df = pd.read_csv(os.path.join(csv_dir, csvname))
            num_cols = df.select_dtypes(include=[np.number]).columns
            df[num_cols] = df[num_cols].round(decimals)
            df.to_excel(writer, sheet_name=sheet, index=False)
    print('[export] style workbook', flush=True)
    wb = load_workbook(excel_path)
    for ws in wb.worksheets:
        style_ws(ws)
    wb.save(excel_path)
    return excel_path


def run_all_steps(excel_path: str, media_path: str, out_dir: str, neo4j_uri: str, neo4j_user: str, neo4j_pass: str, neo4j_db: str, n_splits: int, random_state: int, min_label_freq: int, heat_ablation: bool=True, ppr_heat_teleport_mass: float=0.35, metapath_heat_teleport_mass: float=0.35) -> None:
    ensure_dir(out_dir)
    global EXCEL_PATH, MEDIA_PATH, RANDOM_STATE, N_SPLITS, FE_MIN_LABEL_FREQ_FOR_TABLE
    EXCEL_PATH, MEDIA_PATH = excel_path, media_path
    RANDOM_STATE, N_SPLITS, FE_MIN_LABEL_FREQ_FOR_TABLE = random_state, n_splits, min_label_freq

    print('\n==============================')
    print('[STEP0] Load data and build dataset')
    print('==============================')
    builder = DataBuilder(excel_path, media_path, neo4j_db, out_dir=out_dir)
    df = builder.load()
    df.to_csv(os.path.join(out_dir, 'STEP0_dataset_with_demand_structure.csv'), index=False, encoding='utf-8-sig')
    print('[STEP0b] Write KG reasoning outputs (PPR/MPPR for Table 7 and Appendix E)', flush=True)
    write_kg_reasoning_outputs(df, out_dir=out_dir, media_path=media_path, heat_ablation=heat_ablation, ppr_heat_teleport_mass=ppr_heat_teleport_mass, metapath_heat_teleport_mass=metapath_heat_teleport_mass)
    # stats and splits
    stats = pd.DataFrame([
        {'variable':'price','count':len(df),'mean':df['price'].mean(),'std':df['price'].std(ddof=1),'min':df['price'].min(),'max':df['price'].max()},
        {'variable':'ln(price)','count':len(df),'mean':df['y_ln'].mean(),'std':df['y_ln'].std(ddof=1),'min':df['y_ln'].min(),'max':df['y_ln'].max()},
    ])
    stats.to_csv(os.path.join(out_dir,'STEP0_price_stats.csv'), index=False, encoding='utf-8-sig')
    kf = KFold(n_splits=n_splits, shuffle=True, random_state=random_state)
    folds_k = [{'train_idx':tr.tolist(),'test_idx':te.tolist()} for tr,te in kf.split(df)]
    save_json(os.path.join(out_dir,'STEP0_splits_kfold.json'), folds_k)
    groups = df['supplier'].astype(str).values
    gkf = GroupKFold(n_splits=min(n_splits, len(np.unique(groups))))
    folds_g = [{'train_idx':tr.tolist(),'test_idx':te.tolist()} for tr,te in gkf.split(df, df['y_ln'].values, groups)]
    save_json(os.path.join(out_dir,'STEP0_splits_groupkfold_supplier.json'), folds_g)

    print('\n==============================')
    print('[STEP1] Build manuscript tables')
    print('==============================')
    tabs = build_tables(df, out_dir)
    csv_dir = os.path.join(out_dir, 'table_out', 'paper_tables_csv')
    ensure_dir(csv_dir)
    manifest = {}
    for k, v in tabs.items():
        csvname = f'{k}.csv'
        manuscript_table(k, v).to_csv(os.path.join(csv_dir, csvname), index=False, encoding='utf-8-sig')
        manifest[k.replace('Table','Table_')] = csvname
    save_json(os.path.join(out_dir, 'table_manifest.json'), manifest)
    export_tables_1_to_11_and_appendices_1_to_5(out_dir)
    print('[OK] all tables generated')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--excel', default=EXCEL_PATH)
    ap.add_argument('--media', default=MEDIA_PATH)
    ap.add_argument('--out', default=OUTPUT_PATH)
    ap.add_argument('--neo4j_db', default='./anymous/neo4j_export')
    ap.add_argument('--n_splits', type=int, default=N_SPLITS)
    ap.add_argument('--seed', type=int, default=RANDOM_STATE)
    ap.add_argument('--min_label_freq', type=int, default=FE_MIN_LABEL_FREQ_FOR_TABLE)
    args = ap.parse_args()
    run_all_steps(args.excel, args.media, args.out, '', '', '', args.neo4j_db, args.n_splits, args.seed, args.min_label_freq)

if __name__ == '__main__':
    main()
