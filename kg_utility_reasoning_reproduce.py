#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations
import os, json, math, argparse, importlib.util, subprocess, sys
from typing import Any, Dict, List
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def ensure_dir(p: str) -> None:
    os.makedirs(p, exist_ok=True)


def find_anymous_root(d: str) -> str:
    d = os.path.abspath(d)
    if os.path.exists(os.path.join(d, 'name_price_anonymized.xlsx')):
        return d
    cand = os.path.join(d, 'anymous')
    if os.path.exists(os.path.join(cand, 'name_price_anonymized.xlsx')):
        return cand
    for cur, _, files in os.walk(d):
        if 'name_price_anonymized.xlsx' in files and 'media_result.xlsx' in files and os.path.isdir(os.path.join(cur, 'neo4j_export')):
            return cur
    raise FileNotFoundError(f'Cannot locate valid data root under: {d}')


def load_module(name: str, path: str):
    spec = importlib.util.spec_from_file_location(name, os.path.abspath(path))
    mod = importlib.util.module_from_spec(spec)
    sys.modules[name] = mod
    spec.loader.exec_module(mod)
    return mod


def style_ws(ws):
    header_fill = PatternFill(fill_type='solid', fgColor='D9E2F3')
    panel_fill = PatternFill(fill_type='solid', fgColor='EAF2F8')
    thin = Side(style='thin', color='BFBFBF')
    bottom = Border(bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill
        cell.border = bottom
    ws.freeze_panes = 'A2'
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=2):
        first = row[0].value
        if isinstance(first, str) and first.startswith('Panel '):
            for cell in row[:max_col]:
                cell.font = Font(bold=True, italic=True)
                cell.fill = panel_fill
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = bottom
        else:
            for cell in row[:max_col]:
                cell.alignment = Alignment(horizontal='center' if cell.column > 2 else 'left', vertical='center', wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = '0.000'
    for col in ws.columns:
        mx = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(mx+2, 10), 55)


def parse_list(x: Any) -> List[str]:
    if isinstance(x, list):
        return [str(v).strip() for v in x if str(v).strip()]
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return []
    s = str(x).strip()
    if not s or s.lower() in {'nan','none','null'}:
        return []
    if s.startswith('[') and s.endswith(']'):
        try:
            import ast
            v = ast.literal_eval(s)
            if isinstance(v, list):
                return [str(t).strip() for t in v if str(t).strip()]
        except Exception:
            pass
    return [t.strip() for t in s.split('|') if t.strip()]


def fmt3(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        if pd.api.types.is_numeric_dtype(out[c]):
            out[c] = out[c].round(3)
    return out


def build_appendix_tables(any_root: str, out_dir: str) -> Dict[str, str]:
    neo_dir = os.path.join(any_root, 'neo4j_export')
    nodes_dp = pd.read_csv(os.path.join(neo_dir, 'nodes_dataproduct.csv')).rename(columns={'name_anon':'name','supplier_anon':'supplier','desc_anon':'desc'})
    rel_app = pd.read_csv(os.path.join(neo_dir, 'rel_applied_to.csv')).rename(columns={'app_name':'app'})
    rel_src = pd.read_csv(os.path.join(neo_dir, 'rel_source_industry.csv')).rename(columns={'src_name':'src'})
    price = pd.read_excel(os.path.join(any_root, 'name_price_anonymized.xlsx')).copy()
    price['price'] = pd.to_numeric(price['price'], errors='coerce')
    price = price[price['price'].gt(0)].copy()
    price['name'] = price['name'].astype(str)
    price['supplier'] = price['supplier'].astype(str)
    price = price.drop_duplicates(['name','supplier'])
    dp = nodes_dp[['dp_id','name','supplier']].copy()
    dp['name'] = dp['name'].astype(str)
    dp['supplier'] = dp['supplier'].astype(str)

    # Appendix C1: auditable KG universe and price-modeling sample construction
    raw_price_path = os.path.join(any_root, 'name_price_all_anonymized.xlsx')
    if not os.path.exists(raw_price_path):
        raise FileNotFoundError(
            'Appendix C1 requires raw price file name_price.xlsx under the anonymized root. '
            'This file is needed to reproduce raw rows, zero-price records, and upper-tail exclusions.'
        )

    raw_price = pd.read_excel(raw_price_path).copy()
    raw_price['price'] = pd.to_numeric(raw_price['price'], errors='coerce')
    raw_price['name'] = raw_price['name'].astype(str).str.strip()
    raw_price['supplier'] = raw_price['supplier'].astype(str).str.strip()

    kg_total_products = int(nodes_dp['dp_id'].astype(str).nunique()) if 'dp_id' in nodes_dp.columns else int(nodes_dp[['name','supplier']].drop_duplicates().shape[0])
    raw_rows = int(len(raw_price))
    raw_unique_pairs = int(raw_price[['name','supplier']].drop_duplicates().shape[0])
    zero_price_records = int(raw_price['price'].eq(0).sum())
    upper_tail_removed = int(raw_price['price'].ge(300).sum())
    final_model_sample = int(price[['name','supplier']].drop_duplicates().shape[0])

    # price_dp and price_ids
    priced_dp = price.merge(dp[['dp_id','name','supplier']], on=['name','supplier'], how='inner')
    priced_ids = set(priced_dp['dp_id'].astype(str))

    # Appendix C1 content
    final_matched_to_kg = int(priced_dp[['name','supplier']].drop_duplicates().shape[0])

    c1 = pd.DataFrame([
        {
            'Stage': 'External KG API DataProduct universe',
            'Count': kg_total_products,
            'Audit/source basis': 'Unique API DataProduct nodes in the anonymized Neo4j export after cleaning, de-duplication, anonymization, and KG instantiation.'
        },
        {
            'Stage': 'Raw price-file rows',
            'Count': raw_rows,
            'Audit/source basis': 'Rows in the raw price file used for price cleaning.'
        },
        {
            'Stage': 'Unique product-supplier pairs in raw price file',
            'Count': raw_unique_pairs,
            'Audit/source basis': 'Product name and supplier name are used as the uniqueness criterion.'
        },
        {
            'Stage': 'Zero-price records removed',
            'Count': zero_price_records,
            'Audit/source basis': 'Records with price = 0; excluded because they do not represent positive paid API posted quotes.'
        },
        {
            'Stage': 'Upper-tail observations removed',
            'Count': upper_tail_removed,
            'Audit/source basis': 'Records priced at 300 RMB per call or above.'
        },
        {
            'Stage': 'Final price-modeling sample',
            'Count': final_model_sample,
            'Audit/source basis': 'Unique API products with positive normalized posted prices after cleaning.'
        },
        {
            'Stage': 'Final sample matched back to KG',
            'Count': final_matched_to_kg,
            'Audit/source basis': 'Final modeling records matched to the KG through anonymized product and supplier labels.'
        },
    ])

    # Appendix C2
    def degree_stats(start_ids, end_ids, edges_df, start_col, end_col):
        outdeg = edges_df.groupby(start_col)[end_col].count().reindex(start_ids, fill_value=0)
        indeg = edges_df.groupby(end_col)[start_col].count().reindex(end_ids, fill_value=0)
        def pack(s):
            s = pd.Series(s, dtype=float)
            return f"{s.mean():.3f}; {int(s.quantile(0.25))}/{int(s.quantile(0.5))}/{int(s.quantile(0.75))}; {int(s.max())}"
        return pack(outdeg), pack(indeg)

    provide = priced_dp[['supplier','dp_id']].rename(columns={'supplier':'start','dp_id':'end'})
    src_edges = rel_src[rel_src['dp_id'].astype(str).isin(priced_ids)][['src','dp_id']].rename(columns={'src':'start','dp_id':'end'})
    app_edges = rel_app[rel_app['dp_id'].astype(str).isin(priced_ids)][['dp_id','app']].rename(columns={'dp_id':'start','app':'end'})

    rows_c2 = []
    for rel_name, schema, edf, start_set, end_set in [
        ('provide_data','Supplier → DataProduct', provide, sorted(provide['start'].astype(str).unique()), sorted(provide['end'].astype(str).unique())),
        ('source_industry','src_IndustryCategory → DataProduct', src_edges, sorted(src_edges['start'].astype(str).unique()), sorted(src_edges['end'].astype(str).unique())),
        ('applied_to','DataProduct → app_IndustryCategory', app_edges, sorted(app_edges['start'].astype(str).unique()), sorted(app_edges['end'].astype(str).unique())),
    ]:
        sdeg, edeg = degree_stats(start_set, end_set, edf, 'start', 'end')
        rows_c2.append({
            'Relation (r)': rel_name,
            'Schema (Start → End)': schema,
            '#Edges': len(edf),
            '#Start': len(start_set),
            '#End': len(end_set),
            'Start out-degree (mean; P25/P50/P75; max)': sdeg,
            'End in-degree (mean; P25/P50/P75; max)': edeg,
        })
    c2 = pd.DataFrame(rows_c2)

    # Appendix C3
    app_counts = rel_app[rel_app['dp_id'].astype(str).isin(priced_ids)].groupby('dp_id').size()
    src_counts = rel_src[rel_src['dp_id'].astype(str).isin(priced_ids)].groupby('dp_id').size()
    prod_per_supplier = priced_dp.groupby('supplier').size()
    def pack_stat(s):
        s = pd.Series(s, dtype=float)
        return f"{s.mean():.3f}; {int(s.quantile(0.25))}/{int(s.quantile(0.5))}/{int(s.quantile(0.75))}; {int(s.max())}"
    c3 = pd.DataFrame([
        {'Panel':'A. Node coverage','Statistic':'Products (with price)','Value':len(priced_ids)},
        {'Panel':'A. Node coverage','Statistic':'Suppliers (connected to priced products)','Value':priced_dp['supplier'].nunique()},
        {'Panel':'A. Node coverage','Statistic':'src industries (connected to priced products)','Value':src_edges['start'].nunique()},
        {'Panel':'A. Node coverage','Statistic':'app industries (connected to priced products)','Value':app_edges['end'].nunique()},
        {'Panel':'B. Edge coverage','Statistic':'Triples: provide_data','Value':len(provide)},
        {'Panel':'B. Edge coverage','Statistic':'Triples: source_industry','Value':len(src_edges)},
        {'Panel':'B. Edge coverage','Statistic':'Triples: applied_to','Value':len(app_edges)},
        {'Panel':'C. Degree & label sparsity','Statistic':'Products per supplier (mean; P25/P50/P75; max)','Value':pack_stat(prod_per_supplier)},
        {'Panel':'C. Degree & label sparsity','Statistic':'App labels per product (mean; P25/P50/P75; max)','Value':pack_stat(app_counts.reindex(sorted(priced_ids), fill_value=0))},
        {'Panel':'C. Degree & label sparsity','Statistic':'Src labels per product (mean; P25/P50/P75; max)','Value':pack_stat(src_counts.reindex(sorted(priced_ids), fill_value=0))},
    ])

    # Appendix D1 from the exact STEP0 dataset generated by the core program.
    # This guarantees that Appendix D describes the same paper-facing variables
    # created by the core script. Demand and market variables follow the manuscript/Appendix D definitions.
    step0_path = os.path.join(out_dir, 'STEP0_dataset_with_demand_structure.csv')
    if not os.path.exists(step0_path):
        raise FileNotFoundError(f'Cannot find core STEP0 dataset for Appendix D: {step0_path}')
    d = pd.read_csv(step0_path)

    appendix_d_map = {
        'heat_news': 'heat_news',
        'heat_total': 'heat_total',
        'heat_web': 'heat_web',
        'heat_weixin': 'heat_weixin',
        'Bs': 'Bs',
        'HHI_proxy': 'HHI_proxy',
        'T': 'T',
    }
    missing = [col for col in appendix_d_map.values() if col not in d.columns]
    if missing:
        raise KeyError(f'Missing expected STEP0 variables for Appendix D: {missing}')

    def desc_row(label, col):
        s = pd.to_numeric(d[col], errors='coerce')
        return {
            'Var': label,
            'Obs': int(s.notna().sum()),
            'Mean': s.mean(),
            'SD': s.std(ddof=1),
            'Min': s.min(),
            'P25': s.quantile(0.25),
            'P50': s.quantile(0.5),
            'P75': s.quantile(0.75),
            'Max': s.max(),
        }
    d1 = pd.DataFrame([desc_row(label, col) for label, col in appendix_d_map.items()])
    d1 = fmt3(d1)

    csv_dir = os.path.join(out_dir, 'table_out', 'paper_tables_csv')
    ensure_dir(csv_dir)

    c1_path = os.path.join(csv_dir, 'Appendix_C1.csv')
    c1.to_csv(c1_path, index=False, encoding='utf-8-sig')
    c2_path = os.path.join(csv_dir, 'Appendix_C2.csv');
    fmt3(c2).to_csv(c2_path, index=False)
    c3_path = os.path.join(csv_dir, 'Appendix_C3.csv');
    c3.to_csv(c3_path, index=False)
    d1_path = os.path.join(csv_dir, 'Appendix_D1.csv');
    d1.to_csv(d1_path, index=False)
    return {
        'App_Table_C1':'Appendix_C1.csv',
        'App_Table_C2':'Appendix_C2.csv',
        'App_Table_C3':'Appendix_C3.csv',
        'App_Table_D1':'Appendix_D1.csv'
    }


def merge_excel(out_dir: str, appendix_manifest: Dict[str, str]) -> str:
    csv_dir = os.path.join(out_dir, 'table_out', 'paper_tables_csv')
    lst_csv = os.listdir(csv_dir)
    lst_csv.sort()

    manifest = {}
    for c in lst_csv:
        if c.endswith('.csv'):
            if c.startswith('Appendix'):
                manifest[c.replace('Appendix_', 'Appendix_Table_').strip('.csv')] = c
            else:
                manifest[c.replace('Table', 'Table_').strip('.csv')] = c

    # manifest = json.load(open(os.path.join(out_dir, 'table_manifest.json'), 'r', encoding='utf-8'))

    # Add Appendix C/D outputs generated by the reproduction wrapper.
    # for k, v in appendix_manifest.items():
    #     manifest[k] = v

    # Explicit manuscript-facing order.
    ordered_sheets = [
        'Table_1',
        'Table_2',
        'Table_3',
        'Table_4',
        'Table_5',
        'Table_6',
        'Table_7',
        'Appendix_Table_C1',
        'Appendix_Table_C2',
        'Appendix_Table_C3',
        'Appendix_Table_D1',
        'Appendix_Table_E1',
        'Appendix_Table_E2',
        'Appendix_Table_E3',
    ]

    missing = [s for s in ordered_sheets if s not in manifest]
    if missing:
        raise KeyError(f'Missing tables in merge manifest: {missing}')

    merged_dir = os.path.join(out_dir, 'table_merge')
    ensure_dir(merged_dir)
    merged_path = os.path.join(merged_dir, 'Paper_All_Tables_Merged.xlsx')

    with pd.ExcelWriter(merged_path, engine='openpyxl') as writer:
        for sheet in ordered_sheets:
            csvname = manifest[sheet]
            csv_path = os.path.join(csv_dir, csvname)
            if not os.path.exists(csv_path):
                raise FileNotFoundError(f'Missing CSV for {sheet}: {csv_path}')
            df = pd.read_csv(csv_path)
            for c in df.columns:
                if pd.api.types.is_numeric_dtype(df[c]):
                    df[c] = df[c].round(3)
            df.to_excel(writer, sheet_name=sheet[:31], index=False)

    wb = load_workbook(merged_path)
    for ws in wb.worksheets:
        style_ws(ws)
    wb.save(merged_path)
    return merged_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', default='./anymous')
    ap.add_argument('--core', default='./kg_utility_reasoning.py')
    ap.add_argument('--fig_stats_script', default='./plt_price_distribution.py')
    ap.add_argument('--fig_script', default='./plt_kg_reasoning_pic.py')
    ap.add_argument('--out_dir', default='./result_kg_reproduce')
    ap.add_argument('--min_label_freq', type=int, default=3)
    args = ap.parse_args()

    root = find_anymous_root(args.input)
    print(f'>>> Root data directory determined as: {root}')
    price_xlsx = os.path.join(root, 'name_price_anonymized.xlsx')
    media_xlsx = os.path.join(root, 'media_result.xlsx')
    neo_dir = os.path.join(root, 'neo4j_export')

    core = load_module('kg_core_fixed', args.core)
    print('>>> Loading core reproduction script...')
    core.run_all_steps(
        excel_path=price_xlsx,
        media_path=media_xlsx,
        out_dir=args.out_dir,
        neo4j_uri='', neo4j_user='', neo4j_pass='', neo4j_db=neo_dir,
        n_splits=5,
        random_state=42,
        min_label_freq=args.min_label_freq,
        heat_ablation=True,
        ppr_heat_teleport_mass=0.35,
        metapath_heat_teleport_mass=0.35,
    )

    print('>>> Reproducing Fig. 2...')
    subprocess.run([sys.executable, args.fig_stats_script, '--out_dir', args.out_dir, '--fig_dir', os.path.join(args.out_dir, 'Figs')], check=True)

    print('>>> Reproducing Fig. 6...')
    subprocess.run([sys.executable, args.fig_script, '--out_dir', args.out_dir, '--fig_dir', os.path.join(args.out_dir, 'Figs')], check=True)

    print('>>> Building Appendix C/D tables from anymous data...')
    app_manifest = build_appendix_tables(root, args.out_dir)
    print('>>> Merging workbook...')
    merged = merge_excel(args.out_dir, app_manifest)
    print(f'[Complete] {merged}')

if __name__ == '__main__':
    main()
